# -*- coding: utf-8 -*-
"""
一脚完成：
1) 2412 MHz 主载波 20 MHz 通道功率 (CPOW)
2) 所有谐波 (2f, 3f, ...)，直到频率 <= 18 GHz

谐波部分：
- 每个谐波中心：n * 2412 MHz
- Span: 100 MHz（如需按 FCC 改，可调整）
- RBW / VBW: 1 MHz
- 衰减列表: [25, 20, 15] dB
- 每个 ATT：INIT:IMM;*WAI → CALC:MARK1:MAX → X?/Y?
"""

import socket
import importlib.util
import time
from datetime import datetime
import re
import subprocess
from copy import copy
import tkinter as tk
from tkinter import simpledialog, filedialog, messagebox
from pathlib import Path
from typing import Optional, List, Dict, Any, Tuple
import csv
import os
import sys
from fractions import Fraction
from wifi_bandedge import WIFI_BANDEDGE_COLUMNS, measure_wifi_bandedges, measure_bandedge_side_max


FSV_IP = "192.168.20.151"
FSV_PORT = 5025
SOCKET_TIMEOUT = 10.0
SCPI_DELAY = 0.0
_FSV_INITIALIZED = False

GUI_HOST = "192.168.20.11"
GUI_PORT = 7481
GUI_TIMEOUT = 5.0
USE_GUI_CALIBRATION = True
CMD_DELAY = 1.0
CONNECT_SETTLE_S = 4.0
BT_TXCFG_TO_START_DELAY_S = 2.0
TX_START_STABLE_S = 0.8
FW_SWITCH_DISCONNECT_S = 0.8
FW_SWITCH_SETTLE_S = 3.0
STOP_AFTER_START_TX = False
SIMPLE_GUI_FLOW = False
STOP_AFTER_CALIBRATION = False
AUTO_LAUNCH_GUI = True
GUI_EXE_PATH = r"E:\Altobeam GUI\WiFi6_GUI_20251223\WiFi6_GUI\AtbmWLANFacility_Customer.exe"

INPUT_CSV_SINGLE_BAND = "FCC_test_item_single_band.xlsx"
INPUT_CSV_DULE_BAND = "FCC_test_item_dule_band.xlsx"
INPUT_CSV_DULE_ANTENNA = "FCC_test_item_Dule_Antenna.xlsx"
INPUT_CSV_BT_BLE = "FCC_test_item_BT_BLE.xlsx"
INPUT_CSV_BANDEDGE = "FCC_test_item_Bandedge.xlsx"
OUTPUT_CSV = "FCC_conduction_result.xlsx"
OUTPUT_CSV_BT = "BT_FCC_conduction_result.xlsx"
OUTPUT_CSV_BANDEDGE = "Bandedge_FCC_conduction_result.xlsx"
LOSS_TABLE_PATH = "loss.txt"
LOSS_TABLE_PATH_DULE_ANTENNA = "loss_Dule_Antenna.txt"
CONFIG_DIR_NAME = "config"
RESULT_DIR_NAME = "result"
RESULT_DIR_BANDEDGE_NAME = "result_bandedge"
RESULT_DIR_BT_NAME = "result_bt"

DEFAULT_TX_CONFIG: Dict[str, object] = {
    "CHAN": 1,
    "BW": "20M",
    "OFFSET": 0,
    "MODE": "DSSS",
    "OFDM_MODE": "MM",
    "RATE": "1M",
    "CODING": "BCC",
    "DUTY_CYCLE": 100,
    "PSDU_LEN": 10000,
    "CONNECT_TYPE": "USB",
    "ANTENNA": "",
    "CERTIFICATION_MODE": "",
    "GPIO20": None,
    "GPIO21": None,
    "FIRMWARE_TYPE": "WIFI",
    "TEST_MODE": "TX",
    "PACKET_TYPE": "BT_BLE_1M",
    "PAYLOAD": 0,
    "PAYLOAD_LEN": 37,
}

KEY_ALIASES = {
    "CHAN": "CHAN",
    "CH": "CHAN",
    "BW": "BW",
    "OFFSET": "OFFSET",
    "MODE": "MODE",
    "OFDM MODE": "OFDM_MODE",
    "OFDM_MODE": "OFDM_MODE",
    "RATE": "RATE",
    "CODING": "CODING",
    "DUTY CYCLE": "DUTY_CYCLE",
    "DUTY_CYCLE": "DUTY_CYCLE",
    "PSDU LEN": "PSDU_LEN",
    "PSDU_LEN": "PSDU_LEN",
    "ANTENNA": "ANTENNA",
    "CERTIFICATION": "CERTIFICATION_MODE",
    "CERTIFICATION MODE": "CERTIFICATION_MODE",
    "CERTIFICATION_MODE": "CERTIFICATION_MODE",
    "GPIO20": "GPIO20",
    "GPIO21": "GPIO21",
    "FIRMWARE TYPE": "FIRMWARE_TYPE",
    "FIRMWARE_TYPE": "FIRMWARE_TYPE",
    "TEST MODE": "TEST_MODE",
    "TEST_MODE": "TEST_MODE",
    "PACKET TYPE": "PACKET_TYPE",
    "PACKET_TYPE": "PACKET_TYPE",
    "PACKETTYPE": "PACKET_TYPE",
    "PCKTYPE": "PACKET_TYPE",
    "PAYLOAD": "PAYLOAD",
    "PAYLOAD LEN": "PAYLOAD_LEN",
    "PAYLOAD_LEN": "PAYLOAD_LEN",
    "PAYLOADLEN": "PAYLOAD_LEN",
}

FIRMWARE_TYPES = {"WIFI", "BLE", "WIFI_AND_BLE"}
TEST_MODES = {"TX", "RX"}
BT_PACKET_TYPES = {
    "BT_BLE_1M",
    "BT_BLE_2M",
    "BT_BLE_S8",
    "BT_BLE_S2",
    "BT_ID",
    "BT_NULL",
    "BT_POLL",
    "BT_FHS",
    "BT_DM1",
    "BT_DH1",
    "BT_2_DH1",
    "BT_HV1",
    "BT_HV2",
    "BT_2_EV3",
    "BT_HV3",
    "BT_EV3",
    "BT_3_EV3",
    "BT_DV",
    "BT_3_DH1",
    "BT_AUX1",
    "BT_DM3",
    "BT_2_DH3",
    "BT_DH3",
    "BT_3_DH3",
    "BT_EV4",
    "BT_2_EV5",
    "BT_EV5",
    "BT_3_EV5",
    "BT_DM5",
    "BT_2_DH5",
    "BT_DH5",
    "BT_3_DH5",
}


def _normalize_firmware_type(value: object, default: str = "WIFI") -> str:
    raw = str(value or "").strip().upper()
    if not raw:
        return default
    if raw in {"BT", "BLE"}:
        return "BLE"
    if raw in {"WIFI+BT", "WIFI_AND_BT", "WIFI_AND_BLE"}:
        return "WIFI_AND_BLE"
    if raw == "WIFI":
        return "WIFI"
    return default


def _normalize_bt_packet_type(value: object) -> str:
    raw = str(value or "").strip().upper().replace("-", "_")
    if not raw:
        return ""
    if raw.startswith("BT_"):
        return raw
    if raw.startswith("BLE_"):
        return f"BT_{raw}"
    return raw

F0_HZ_DEFAULT = 2.412e9      # 默认 2412 MHz 基波
MAX_FREQ_HZ = 18e9           # 谐波最高测到 18 GHz


class FsvSocket:
    def __init__(self, ip: str, port: int, timeout: float = 10.0):
        self.ip = ip
        self.port = port
        self.timeout = timeout
        self.sock: Optional[socket.socket] = None

    def connect(self):
        s = socket.socket(socket.AF_INET, socket.SOCK_STREAM)
        s.settimeout(self.timeout)
        print(f"[INFO] Connecting to {self.ip}:{self.port} ...")
        s.connect((self.ip, self.port))
        self.sock = s
        print("[INFO] Connected.")

    def close(self):
        if self.sock:
            try:
                self.sock.close()
            except Exception:
                pass
            self.sock = None
            print("[INFO] Socket closed.")

    def send_cmd(self, cmd: str, read_reply: bool = False, bufsize: int = 8192):
        if not self.sock:
            raise RuntimeError("Socket not connected")
        msg = (cmd + "\n").encode("ascii")
        print("SCPI >>", cmd)
        self.sock.sendall(msg)
        if not read_reply:
            _sleep_cmd(SCPI_DELAY)
            return None
        reply = self.sock.recv(bufsize)
        text = reply.decode(errors="ignore").strip()
        print("SCPI <<", text)
        _sleep_cmd(SCPI_DELAY)
        return text

    def query(self, cmd: str, bufsize: int = 8192) -> str:
        return self.send_cmd(cmd, read_reply=True, bufsize=bufsize) or ""

    def query_float(self, cmd: str, bufsize: int = 8192) -> float:
        return float(self.query(cmd, bufsize=bufsize))

    def check_error(self, label: str = "") -> str:
        err = self.query("SYST:ERR?")
        if not err.startswith("0"):
            print(f"[SCPI ERROR]{' '+label if label else ''} {err}")
        return err


def _ensure_fsv_initialized(inst: FsvSocket) -> None:
    global _FSV_INITIALIZED
    if _FSV_INITIALIZED:
        return
    inst.send_cmd(r"MMEM:LOAD:STAT 1,'C:\R_S\Instr\user/QuickSave\QuickSave8'")
    inst.send_cmd("*RST")
    _FSV_INITIALIZED = True


def measure_cpow_20m(
    inst: FsvSocket,
    f0_hz: float,
    loss_table: Optional[List[Tuple[float, float, float]]] = None,
    channel_bw_hz: float = 20e6,
) -> float:
    """
    使用 CPOW 读取主载波 20 MHz 通道功率（dBm）
    只在首次初始化时做 *RST，后续不再重复。
    """
    print("\n===== Step1: 2412 MHz 20MHz 通道功率 (CPOW) =====")

    _ensure_fsv_initialized(inst)
    inst.send_cmd("INIT:CONT OFF")
    inst.send_cmd(f"SENS:FREQ:CENT {f0_hz:.0f}")
    inst.send_cmd("SENS:FREQ:SPAN 100000000")  # 100 MHz
    inst.query("SENS:FREQ:SPAN?")

    inst.send_cmd("DISP:WIND:TRAC:Y:SCAL:RLEV 30")
    ref_offs = 10.0
    if loss_table:
        loss_db = _lookup_cable_loss_db(f0_hz, loss_table)
        if loss_db is not None:
            ref_offs = loss_db
    inst.send_cmd(f"DISP:WIND:TRAC:Y:SCAL:RLEV:OFFS {ref_offs}")
    inst.send_cmd("INP:ATT:AUTO OFF")
    inst.send_cmd("INP:ATT 25")

    # RBW =1MHZ VBW = 3MHz
    inst.send_cmd("SENS:BAND:AUTO OFF")
    inst.send_cmd("SENS:BAND 1MHz")
    inst.send_cmd("SENS:BAND:VID:AUTO OFF")
    inst.send_cmd("SENS:BAND:VID 3MHz")

    # Trace / 平均（跟你之前 step1 一致）
    inst.send_cmd("DISP:WIND:SUBW:TRAC1:MODE AVER")
    inst.send_cmd("SENS:WIND:DET1:FUNC RMS")
    inst.send_cmd("SENS:AVER:TYPE POW")
    inst.send_cmd("SENS:AVER:COUN 100")

    # 20 MHz 通道带宽
    inst.send_cmd(f"SENS:POW:ACH:BWID:CHAN1 {int(channel_bw_hz)}")

    # CPOW 功能
    inst.send_cmd("CALC:MARK:FUNC:POW:SEL CPOW")
    inst.send_cmd("SENS:FREQ:SPAN 100000000")
    inst.query("SENS:FREQ:SPAN?")

    inst.send_cmd("INIT")
    inst.send_cmd("*WAI")

    cpow_str = inst.query("CALC:MARK:FUNC:POW:RES? CPOW")
    first_field = cpow_str.split(",")[0].strip()
    cpow = float(first_field)
    print(f"[INFO] Integrated channel bandwidth: {channel_bw_hz/1e6:.0f} MHz")

    print(f"[RESULT] 主载波 {f0_hz/1e9:.4f} GHz, 20MHz 通道功率 = {cpow:.2f} dBm")

    inst.check_error("after CPOW")

    return cpow


def _load_gui_client_class():
    gui_path = Path(__file__).parent / "GUI control" / "GUI_control.py"
    spec = importlib.util.spec_from_file_location("gui_control", str(gui_path))
    if not spec or not spec.loader:
        raise RuntimeError("Failed to load GUI_control.py")
    module = importlib.util.module_from_spec(spec)
    spec.loader.exec_module(module)
    return module.GuiSocketClient

def _parse_power_value(resp: str) -> float:
    if not resp:
        raise ValueError("Empty POWER_GET response")
    cleaned = resp.replace(",", " ").strip()
    for token in reversed(cleaned.split()):
        try:
            return float(token)
        except ValueError:
            continue
    raise ValueError(f"Unable to parse POWER_GET response: {resp!r}")

def _sleep_cmd(delay_s: float) -> None:
    if delay_s > 0:
        time.sleep(delay_s)

def measure_cpow_with_power_calibration(
    inst: FsvSocket,
    gui,
    f0_hz: float,
    tolerance_db: float = 0.5,
    step_db: float = 0.25,
    max_iters: int = 20,
    control_tx: bool = True,
    desired_target: Optional[float] = None,
    cmd_delay: float = 0.0,
    loss_table: Optional[List[Tuple[float, float, float]]] = None,
    channel_bw_hz: float = 20e6,
) -> Tuple[float, float, float]:
    """
    Sequence: POWER_GET -> START_TX -> CPOW measure -> optional calibration -> STOP_TX.
    Returns (Power_Target, Pwr, Pwr_Tar).
    """
    resp = gui.power_get()
    _sleep_cmd(cmd_delay)
    if desired_target is None:
        desired_target = _parse_power_value(resp)
    if desired_target is None:
        raise ValueError("Missing desired power target")
    current_target = desired_target
    gui.power_target(current_target)
    _sleep_cmd(cmd_delay)
    if control_tx:
        gui.start_tx()
        _sleep_cmd(cmd_delay)
        if TX_START_STABLE_S > 0:
            time.sleep(TX_START_STABLE_S)
    try:
        cpow = measure_cpow_20m(inst, f0_hz, loss_table=loss_table, channel_bw_hz=channel_bw_hz)
        if abs(cpow - desired_target) > tolerance_db:
            # First jump: adjust by full delta, then fine-tune with 0.25 dB steps.
            current_target = desired_target + (desired_target - cpow)
            current_target = round(current_target / step_db) * step_db
            gui.power_target(current_target)
            _sleep_cmd(cmd_delay)
            cpow = measure_cpow_20m(inst, f0_hz, loss_table=loss_table, channel_bw_hz=channel_bw_hz)
            if abs(cpow - desired_target) > tolerance_db:
                for _ in range(max_iters):
                    if cpow < desired_target:
                        current_target += step_db
                    else:
                        current_target -= step_db
                    current_target = round(current_target / step_db) * step_db
                    gui.power_target(current_target)
                    _sleep_cmd(cmd_delay)
                    cpow = measure_cpow_20m(inst, f0_hz, loss_table=loss_table, channel_bw_hz=channel_bw_hz)
                    if abs(cpow - desired_target) <= tolerance_db:
                        break
        final_target = _parse_power_value(gui.power_get())
        _sleep_cmd(cmd_delay)
        return desired_target, cpow, final_target
    finally:
        if control_tx:
            gui.stop_tx()
            _sleep_cmd(cmd_delay)

def _normalize_key(key: str) -> Optional[str]:
    cleaned = key.strip().upper().replace("_", " ")
    return KEY_ALIASES.get(cleaned)

def _parse_chan(value: str) -> Optional[int]:
    raw = value.strip()
    if not raw:
        return None
    upper = raw.upper()
    if upper.startswith("CH"):
        raw = raw[2:]
    try:
        return int(float(raw))
    except ValueError:
        return None


def _parse_wifi_channel_bw_hz(value: object, default_hz: float = 20e6) -> float:
    raw = str(value or "").strip().upper()
    if not raw:
        return default_hz
    cleaned = raw.replace("MHZ", "M").replace(" ", "")
    if cleaned.endswith("M"):
        num = cleaned[:-1]
        try:
            bw_mhz = float(num)
            if bw_mhz > 0:
                return bw_mhz * 1e6
        except Exception:
            return default_hz
        return default_hz
    try:
        numeric = float(cleaned)
        if numeric > 1e6:
            return numeric
        if numeric > 0:
            return numeric * 1e6
    except Exception:
        return default_hz
    return default_hz

def _parse_gpio_value(value: str) -> Optional[int]:
    raw = value.strip().upper()
    if not raw:
        return None
    if raw in {"LOW", "L"}:
        return 0
    if raw in {"HIGH", "H"}:
        return 1
    try:
        return int(float(raw))
    except ValueError:
        return None

def _coerce_gpio_int(value: object) -> Optional[int]:
    if value is None:
        return None
    if isinstance(value, (int, float)):
        return 1 if int(value) else 0
    if isinstance(value, str):
        return _parse_gpio_value(value)
    return None

def _normalize_gpio_level(value: object) -> Optional[str]:
    if value is None:
        return None
    if isinstance(value, str):
        parsed = _parse_gpio_value(value)
        if parsed is None:
            return None
        return "HIGH" if parsed else "LOW"
    if isinstance(value, (int, float)):
        return "HIGH" if int(value) else "LOW"
    return None

def _should_skip_calibration(antenna: str, gpio20: object, gpio21: object) -> bool:
    ant = antenna.strip().upper()
    g20 = _coerce_gpio_int(gpio20)
    g21 = _coerce_gpio_int(gpio21)
    if g20 is None or g21 is None:
        return False
    if ant == "ANT1" and g20 == 0 and g21 == 1:
        return True
    if ant == "ANT2" and g20 == 1 and g21 == 0:
        return True
    return False

def _parse_config_cell(cell: str, out: Dict[str, object]) -> None:
    if ":" not in cell and "：" not in cell:
        return
    if ":" in cell:
        key, value = cell.split(":", 1)
    else:
        key, value = cell.split("：", 1)
    norm = _normalize_key(key)
    if not norm:
        return
    val = value.strip()
    if not val:
        return
    if norm == "CHAN":
        chan = _parse_chan(val)
        if chan is not None:
            out[norm] = chan
        return
    if norm == "ANTENNA":
        out[norm] = val.upper()
        return
    if norm == "CERTIFICATION_MODE":
        out[norm] = val.upper()
        return
    if norm == "FIRMWARE_TYPE":
        out[norm] = _normalize_firmware_type(val, default="WIFI")
        return
    if norm == "TEST_MODE":
        out[norm] = val.upper()
        return
    if norm == "PACKET_TYPE":
        out[norm] = _normalize_bt_packet_type(val)
        return
    if norm == "OFFSET":
        try:
            out[norm] = float(val)
        except ValueError:
            pass
        return
    if norm in {"DUTY_CYCLE", "PSDU_LEN", "PAYLOAD", "PAYLOAD_LEN"}:
        try:
            out[norm] = int(float(val))
        except ValueError:
            pass
        return
    if norm in {"GPIO20", "GPIO21"}:
        out[norm] = _parse_gpio_value(val)
        return
    out[norm] = val

def _is_config_header_row(row: List[str]) -> bool:
    for cell in row:
        if cell is None:
            continue
        cell_str = str(cell).strip()
        if not cell_str:
            continue
        if ":" in cell_str:
            return False
        if cell_str.upper() == "CH":
            return False
    normalized = 0
    for cell in row:
        if cell is None:
            continue
        cell_str = str(cell).strip()
        if not cell_str:
            continue
        if _normalize_key(cell_str):
            normalized += 1
    return normalized >= 2

def _apply_config_header_row(header: List[str], row: List[str], out: Dict[str, object]) -> None:
    for i, key in enumerate(header):
        if key is None:
            continue
        key_str = str(key).strip()
        if not key_str:
            continue
        norm = _normalize_key(key_str)
        if not norm:
            continue
        val = row[i] if i < len(row) else ""
        if val is None:
            continue
        val_str = str(val).strip()
        if not val_str:
            continue
        if norm == "CHAN":
            chan = _parse_chan(val_str)
            if chan is not None:
                out[norm] = chan
            continue
        if norm == "OFFSET":
            try:
                out[norm] = float(val_str)
            except ValueError:
                pass
            continue
        if norm in {"DUTY_CYCLE", "PSDU_LEN", "PAYLOAD", "PAYLOAD_LEN"}:
            try:
                out[norm] = int(float(val_str))
            except ValueError:
                pass
            continue
        if norm in {"FIRMWARE_TYPE", "TEST_MODE", "PACKET_TYPE"}:
            if norm == "FIRMWARE_TYPE":
                out[norm] = _normalize_firmware_type(val_str, default="WIFI")
            elif norm == "PACKET_TYPE":
                out[norm] = _normalize_bt_packet_type(val_str)
            else:
                out[norm] = val_str.upper()
            continue
        if norm in {"GPIO20", "GPIO21"}:
            out[norm] = _parse_gpio_value(val_str)
            continue
        out[norm] = val_str

def _normalize_connect_type(value: str) -> Optional[str]:
    cleaned = value.strip().upper()
    return cleaned or None


def _prompt_dut_name() -> str:
    try:
        root = tk.Tk()
        root.withdraw()
        root.attributes("-topmost", True)
        name = simpledialog.askstring("Input DUT", "Input DUT name (e.g. oceanus):")
        root.destroy()
        return (name or "").strip()
    except Exception:
        try:
            return input("Input DUT name (e.g. oceanus): ").strip()
        except Exception:
            return ""


def _prompt_test_profile() -> str:
    try:
        root = tk.Tk()
        root.withdraw()
        root.attributes("-topmost", True)
        choice_var = tk.StringVar(value="SINGLE_BAND")
        dialog = tk.Toplevel(root)
        dialog.title("Select Profile")
        dialog.attributes("-topmost", True)
        dialog.resizable(False, False)

        tk.Label(dialog, text="Select test profile:").pack(padx=12, pady=(12, 4))
        tk.Radiobutton(dialog, text="SINGLE_BAND", variable=choice_var, value="SINGLE_BAND").pack(
            anchor="w", padx=12
        )
        tk.Radiobutton(dialog, text="DULE_BAND", variable=choice_var, value="DULE_BAND").pack(
            anchor="w", padx=12
        )
        tk.Radiobutton(dialog, text="Dule_Antenna", variable=choice_var, value="Dule_Antenna").pack(
            anchor="w", padx=12
        )

        result = {"value": None}

        def on_ok():
            result["value"] = choice_var.get()
            dialog.destroy()

        def on_cancel():
            result["value"] = None
            dialog.destroy()

        btn_frame = tk.Frame(dialog)
        btn_frame.pack(pady=12)
        tk.Button(btn_frame, text="OK", width=8, command=on_ok).pack(side="left", padx=6)
        tk.Button(btn_frame, text="Cancel", width=8, command=on_cancel).pack(
            side="left", padx=6
        )

        dialog.protocol("WM_DELETE_WINDOW", on_cancel)
        dialog.grab_set()
        root.wait_window(dialog)
        root.destroy()
        return (result["value"] or "SINGLE_BAND").strip()
    except Exception:
        try:
            raw = input("Select test profile (single_band/dule_band/dule_antenna): ").strip().lower()
            if raw in {"dule_antenna", "duleantenna"}:
                return "Dule_Antenna"
            if raw in {"dule_band", "duleband"}:
                return "DULE_BAND"
            return "SINGLE_BAND"
        except Exception:
            return "SINGLE_BAND"

def _prompt_user_inputs() -> Tuple[
    str,
    str,
    Optional[float],
    Optional[float],
    Optional[float],
    bool,
    bool,
    bool,
    bool,
    str,
    str,
    Optional[str],
]:
    try:
        root = tk.Tk()
        root.withdraw()
        root.attributes("-topmost", True)
        dialog = tk.Toplevel(root)
        dialog.title("Input DUT")
        dialog.attributes("-topmost", True)
        dialog.resizable(False, False)

        tk.Label(dialog, text="DUT name:").grid(row=0, column=0, sticky="w", padx=12, pady=(12, 4))
        last_dut_name = _load_last_dut_name()
        name_var = tk.StringVar(value=last_dut_name)
        name_entry = tk.Entry(dialog, textvariable=name_var, width=28)
        name_entry.grid(row=1, column=0, columnspan=4, sticky="we", padx=12, pady=(0, 8))

        tk.Label(dialog, text="CONNECT TYPE:").grid(row=2, column=0, sticky="w", padx=12, pady=(4, 4))
        connect_type_var = tk.StringVar(value="USB")
        tk.OptionMenu(dialog, connect_type_var, "USB", "I2C").grid(
            row=2, column=1, columnspan=3, sticky="w", padx=12, pady=(4, 4)
        )

        tk.Label(dialog, text="GUI ADDRESS:").grid(row=3, column=0, sticky="w", padx=12, pady=(4, 4))
        gui_address_var = tk.StringVar(value=GUI_EXE_PATH)
        gui_address_entry = tk.Entry(dialog, textvariable=gui_address_var, width=44)
        gui_address_entry.grid(row=3, column=1, columnspan=2, sticky="we", padx=12, pady=(4, 4))

        def _browse_gui_address() -> None:
            selected = filedialog.askopenfilename(
                parent=dialog,
                title="Select GUI EXE",
                filetypes=[("Executable", "*.exe"), ("All files", "*.*")],
            )
            if selected:
                gui_address_var.set(selected)

        tk.Button(dialog, text="Browse...", width=10, command=_browse_gui_address).grid(
            row=3, column=3, sticky="w", padx=(0, 12), pady=(4, 4)
        )

        tk.Label(dialog, text="Firmware type:").grid(row=4, column=0, sticky="w", padx=12, pady=(4, 4))
        firmware_type_var = tk.StringVar(value="WIFI")
        firmware_wifi_rb = tk.Radiobutton(dialog, text="WIFI", variable=firmware_type_var, value="WIFI")
        firmware_wifi_rb.grid(
            row=4, column=1, sticky="w", padx=12
        )
        firmware_bt_rb = tk.Radiobutton(dialog, text="BT", variable=firmware_type_var, value="BLE")
        firmware_bt_rb.grid(
            row=4, column=2, sticky="w", padx=12
        )
        firmware_both_rb = tk.Radiobutton(
            dialog, text="WIFI+BT", variable=firmware_type_var, value="WIFI_AND_BLE"
        )
        firmware_both_rb.grid(
            row=4, column=3, sticky="w", padx=12
        )

        tk.Label(dialog, text="Profile:").grid(row=5, column=0, sticky="w", padx=12, pady=(4, 4))
        profile_var = tk.StringVar(value="SINGLE_BAND")
        profile_single_rb = tk.Radiobutton(dialog, text="SINGLE_BAND", variable=profile_var, value="SINGLE_BAND")
        profile_single_rb.grid(
            row=6, column=0, sticky="w", padx=12
        )
        profile_dule_band_rb = tk.Radiobutton(dialog, text="DULE_BAND", variable=profile_var, value="DULE_BAND")
        profile_dule_band_rb.grid(
            row=6, column=1, sticky="w", padx=12
        )
        profile_dule_ant_rb = tk.Radiobutton(
            dialog, text="Dule_Antenna", variable=profile_var, value="Dule_Antenna"
        )
        profile_dule_ant_rb.grid(
            row=6, column=2, sticky="w", padx=12
        )

        tk.Label(dialog, text="Band:").grid(row=7, column=0, sticky="w", padx=12, pady=(6, 4))
        band_24_var = tk.BooleanVar(value=True)
        band_5_var = tk.BooleanVar(value=True)
        band_24_cb = tk.Checkbutton(dialog, text="2.4G", variable=band_24_var)
        band_24_cb.grid(row=7, column=1, sticky="w", padx=6)
        band_5_cb = tk.Checkbutton(dialog, text="5G", variable=band_5_var)
        band_5_cb.grid(row=7, column=2, sticky="w", padx=6)

        tk.Label(dialog, text="Test:").grid(row=8, column=0, sticky="w", padx=12, pady=(6, 4))
        harmonic_test_var = tk.BooleanVar(value=True)
        bandedge_test_var = tk.BooleanVar(value=True)
        harmonic_test_cb = tk.Checkbutton(dialog, text="Harmonic", variable=harmonic_test_var)
        harmonic_test_cb.grid(
            row=8, column=1, sticky="w", padx=6
        )
        bandedge_test_cb = tk.Checkbutton(dialog, text="Bandedge", variable=bandedge_test_var)
        bandedge_test_cb.grid(
            row=8, column=2, sticky="w", padx=6
        )

        enable_var = tk.BooleanVar(value=False)
        cal_scope_cb = tk.Checkbutton(dialog, text="Cal Power Scope", variable=enable_var)
        cal_scope_cb.grid(
            row=9, column=0, columnspan=4, sticky="w", padx=12, pady=(8, 4)
        )

        tk.Label(dialog, text="Min").grid(row=10, column=0, sticky="w", padx=12)
        min_var = tk.StringVar(value="-1")
        min_entry = tk.Entry(dialog, textvariable=min_var, width=8)
        min_entry.grid(row=10, column=1, padx=6, pady=4)

        tk.Label(dialog, text="Max").grid(row=10, column=2, sticky="w", padx=6)
        max_var = tk.StringVar(value="0")
        max_entry = tk.Entry(dialog, textvariable=max_var, width=8)
        max_entry.grid(row=10, column=3, padx=6, pady=4)

        tk.Label(dialog, text="Step").grid(row=11, column=0, sticky="w", padx=12)
        step_var = tk.StringVar(value="1")
        step_entry = tk.Entry(dialog, textvariable=step_var, width=8)
        step_entry.grid(row=11, column=1, padx=6, pady=4)

        tip_text = (
            "SWITCH connection:\n"
            "PIN7->GPIO20, PIN8->GPIO21,\n"
            "SMA1->ANT1, SMA2->FSV,\n"
            "SMA3->ANT2, SMA4->50ohm"
        )
        tk.Label(dialog, text=tip_text, justify="left").grid(
            row=12, column=0, columnspan=4, sticky="w", padx=12, pady=(6, 4)
        )

        def _set_scope_state(enabled: bool) -> None:
            state = "normal" if enabled else "disabled"
            min_entry.configure(state=state)
            max_entry.configure(state=state)
            step_entry.configure(state=state)

        def _refresh_ui_state() -> None:
            fw = firmware_type_var.get().strip().upper()
            wifi_path_enabled = fw in {"WIFI", "WIFI_AND_BLE"}
            normal_or_disabled = "normal" if wifi_path_enabled else "disabled"

            profile_single_rb.configure(state=normal_or_disabled)
            profile_dule_band_rb.configure(state=normal_or_disabled)
            profile_dule_ant_rb.configure(state=normal_or_disabled)
            harmonic_test_cb.configure(state=normal_or_disabled)
            bandedge_test_cb.configure(state=normal_or_disabled)
            cal_scope_cb.configure(state=normal_or_disabled)

            if not wifi_path_enabled:
                band_24_var.set(True)
                band_5_var.set(True)
                harmonic_test_var.set(False)
                bandedge_test_var.set(False)
                if enable_var.get():
                    enable_var.set(False)
                band_24_cb.configure(state="disabled")
                band_5_cb.configure(state="disabled")
                _set_scope_state(False)
                return

            if profile_var.get() == "SINGLE_BAND":
                band_24_var.set(True)
                band_5_var.set(False)
                band_24_cb.configure(state="disabled")
                band_5_cb.configure(state="disabled")
            else:
                band_24_cb.configure(state="normal")
                band_5_cb.configure(state="normal")

            scope_enabled = bool(harmonic_test_var.get())
            if not scope_enabled and enable_var.get():
                enable_var.set(False)
            cal_scope_cb.configure(state="normal" if scope_enabled else "disabled")
            _set_scope_state(scope_enabled and bool(enable_var.get()))

        _set_scope_state(False)
        enable_var.trace_add("write", lambda *_: _refresh_ui_state())
        profile_var.trace_add("write", lambda *_: _refresh_ui_state())
        firmware_type_var.trace_add("write", lambda *_: _refresh_ui_state())
        harmonic_test_var.trace_add("write", lambda *_: _refresh_ui_state())
        bandedge_test_var.trace_add("write", lambda *_: _refresh_ui_state())
        _refresh_ui_state()

        result = {"value": None}

        def _parse_value(text: str) -> Optional[float]:
            raw = text.strip()
            if not raw:
                return None
            try:
                return float(raw)
            except Exception:
                return None

        def on_ok():
            if enable_var.get():
                scope_min = _parse_value(min_var.get())
                scope_max = _parse_value(max_var.get())
                scope_step = _parse_value(step_var.get())
            else:
                scope_min = None
                scope_max = None
                scope_step = None
            dut_name = (name_var.get() or "").strip()
            _save_last_dut_name(dut_name)
            selected_gui_address = (gui_address_var.get() or "").strip() or GUI_EXE_PATH
            result["value"] = (
                dut_name,
                profile_var.get().strip().upper(),
                scope_min,
                scope_max,
                scope_step,
                bool(band_24_var.get()),
                bool(band_5_var.get()),
                bool(harmonic_test_var.get()),
                bool(bandedge_test_var.get()),
                connect_type_var.get().strip().upper(),
                firmware_type_var.get().strip().upper(),
                selected_gui_address,
            )
            dialog.destroy()

        def on_cancel():
            result["value"] = None
            dialog.destroy()

        btn_frame = tk.Frame(dialog)
        btn_frame.grid(row=13, column=0, columnspan=4, pady=12)
        tk.Button(btn_frame, text="OK", width=8, command=on_ok).pack(side="left", padx=6)
        tk.Button(btn_frame, text="Cancel", width=8, command=on_cancel).pack(side="left", padx=6)

        dialog.protocol("WM_DELETE_WINDOW", on_cancel)
        dialog.grab_set()
        name_entry.focus_set()
        root.wait_window(dialog)
        root.destroy()
        if result["value"] is None:
            sys.exit(0)
        return result["value"]
    except Exception:
        last_dut_name = _load_last_dut_name()
        try:
            prompt = "Input DUT name (e.g. oceanus)"
            if last_dut_name:
                prompt = f"{prompt} [{last_dut_name}]"
            raw_name = input(f"{prompt}: ").strip()
            name = raw_name or last_dut_name
        except Exception:
            name = last_dut_name or ""
        _save_last_dut_name(name)
        try:
            connect_type = input("CONNECT TYPE (USB/I2C) [USB]: ").strip().upper()
            if connect_type not in {"USB", "I2C"}:
                connect_type = "USB"
        except Exception:
            connect_type = "USB"
        try:
            gui_override = input(f"GUI ADDRESS [{GUI_EXE_PATH}]: ").strip() or GUI_EXE_PATH
        except Exception:
            gui_override = GUI_EXE_PATH
        try:
            raw_fw = input("Firmware type (WIFI/BT/BOTH) [WIFI]: ").strip().upper()
            if raw_fw == "BT":
                firmware_type = "BLE"
            elif raw_fw in {"BOTH", "WIFI+BT", "WIFI_AND_BT", "WIFI_AND_BLE"}:
                firmware_type = "WIFI_AND_BLE"
            else:
                firmware_type = "WIFI"
        except Exception:
            firmware_type = "WIFI"
        wifi_path_enabled = firmware_type in {"WIFI", "WIFI_AND_BLE"}
        profile = "SINGLE_BAND"
        band_24 = True
        band_5 = True
        test_harmonic = False
        test_bandedge = False
        enable = False
        if wifi_path_enabled:
            try:
                raw = input("Select test profile (single_band/dule_band/dule_antenna): ").strip().lower()
                if raw in {"dule_antenna", "duleantenna"}:
                    profile = "Dule_Antenna"
                elif raw in {"dule_band", "duleband"}:
                    profile = "DULE_BAND"
                else:
                    profile = "SINGLE_BAND"
            except Exception:
                profile = "SINGLE_BAND"
            try:
                if profile == "SINGLE_BAND":
                    band_24 = True
                    band_5 = False
                else:
                    band_24 = input("Enable 2.4G? (y/n): ").strip().lower() in {"y", "yes"}
                    band_5 = input("Enable 5G? (y/n): ").strip().lower() in {"y", "yes"}
            except Exception:
                band_24 = True
                band_5 = True
            try:
                harmonic_test = input("Enable Harmonic test? (y/n) [y]: ").strip().lower()
                test_harmonic = harmonic_test not in {"n", "no"}
            except Exception:
                test_harmonic = True
            try:
                bandedge_test = input("Enable Bandedge test? (y/n) [y]: ").strip().lower()
                test_bandedge = bandedge_test not in {"n", "no"}
            except Exception:
                test_bandedge = True
            if test_harmonic:
                try:
                    enable_text = input("Enable Cal Power Scope? (y/n): ").strip().lower()
                    enable = enable_text in {"y", "yes"}
                except Exception:
                    enable = False
        if not enable:
            return (
                name,
                profile,
                None,
                None,
                None,
                band_24,
                band_5,
                test_harmonic,
                test_bandedge,
                connect_type,
                firmware_type,
                gui_override,
            )
        try:
            scope_text = input("Cal Power Scope (min max, e.g. -1 0): ").strip()
            parts = scope_text.replace(",", " ").split()
            scope_min = float(parts[0]) if len(parts) > 0 else None
            scope_max = float(parts[1]) if len(parts) > 1 else None
        except Exception:
            scope_min = None
            scope_max = None
        try:
            step_text = input("Cal Power Step (e.g. 1): ").strip()
            scope_step = float(step_text) if step_text else None
        except Exception:
            scope_step = None
        return (
            name,
            profile,
            scope_min,
            scope_max,
            scope_step,
            band_24,
            band_5,
            test_harmonic,
            test_bandedge,
            connect_type,
            firmware_type,
            gui_override,
        )


def _overwrite_gui_settings_in_file(path: Path, gui_address: str, gui_host: Optional[str] = None) -> bool:
    ext = path.suffix.lower()
    target = gui_address.strip()
    host_target = (gui_host or "").strip()
    if not target and not host_target:
        return False
    target_map: Dict[str, str] = {}
    if target:
        target_map["GUI ADDRESS"] = target
    if host_target:
        target_map["GUI HOST"] = host_target

    if ext == ".xlsx":
        try:
            from openpyxl import load_workbook
        except Exception as exc:
            raise RuntimeError("openpyxl is required to update .xlsx files") from exc
        wb = load_workbook(str(path))
        try:
            ws = wb.active
            changed = False

            non_empty_rows: List[int] = []
            for row_idx in range(1, ws.max_row + 1):
                has_value = False
                for col_idx in range(1, ws.max_column + 1):
                    cell_val = ws.cell(row=row_idx, column=col_idx).value
                    if cell_val is not None and str(cell_val).strip():
                        has_value = True
                        break
                if has_value:
                    non_empty_rows.append(row_idx)
                    if len(non_empty_rows) >= 2:
                        break

            if len(non_empty_rows) >= 2:
                header_row_idx, value_row_idx = non_empty_rows[0], non_empty_rows[1]
                for col_idx in range(1, ws.max_column + 1):
                    key = ws.cell(row=header_row_idx, column=col_idx).value
                    key_clean = str(key or "").strip().upper().replace("_", " ")
                    if key_clean not in target_map:
                        continue
                    old_val = str(ws.cell(row=value_row_idx, column=col_idx).value or "").strip()
                    new_val = target_map[key_clean]
                    if old_val != new_val:
                        ws.cell(row=value_row_idx, column=col_idx).value = new_val
                        changed = True

            for row_idx in range(1, ws.max_row + 1):
                for col_idx in range(1, ws.max_column + 1):
                    cell = ws.cell(row=row_idx, column=col_idx)
                    raw = cell.value
                    if raw is None:
                        continue
                    text = str(raw).strip()
                    if ":" not in text:
                        continue
                    key, val = text.split(":", 1)
                    key_clean = key.strip().upper().replace("_", " ")
                    if key_clean not in target_map:
                        continue
                    new_val = target_map[key_clean]
                    if val.strip() == new_val:
                        continue
                    cell.value = f"{key.strip()}: {new_val}"
                    changed = True

            if changed:
                wb.save(str(path))
            return changed
        finally:
            try:
                wb.close()
            except Exception:
                pass

    if ext == ".csv":
        rows = _read_table_rows(str(path))
        changed = False
        header = None
        values = None
        for row in rows:
            if not row or all(not str(c).strip() for c in row):
                continue
            if header is None:
                header = row
                continue
            values = row
            break
        if header and values:
            for i, key in enumerate(header):
                key_clean = str(key or "").strip().upper().replace("_", " ")
                if key_clean not in target_map:
                    continue
                while i >= len(values):
                    values.append("")
                new_val = target_map[key_clean]
                if str(values[i] or "").strip() != new_val:
                    values[i] = new_val
                    changed = True
        for row in rows:
            for i, cell in enumerate(row):
                text = str(cell or "").strip()
                if ":" not in text:
                    continue
                key, val = text.split(":", 1)
                key_clean = key.strip().upper().replace("_", " ")
                if key_clean not in target_map:
                    continue
                new_val = target_map[key_clean]
                if val.strip() == new_val:
                    continue
                row[i] = f"{key.strip()}: {new_val}"
                changed = True
        if changed:
            _write_table_rows(str(path), rows)
        return changed

    return False


def _override_all_config_gui_address(gui_address: str) -> List[Path]:
    target = gui_address.strip()
    if not target:
        return []
    config_dir = _get_config_dir()
    if not config_dir.exists():
        return []
    changed_files: List[Path] = []
    candidates = sorted(config_dir.glob("*.xlsx")) + sorted(config_dir.glob("*.csv"))
    for cfg_path in candidates:
        try:
            if _overwrite_gui_settings_in_file(cfg_path, target):
                changed_files.append(cfg_path)
        except Exception as exc:
            print(f"[WARN] Failed to update GUI ADDRESS in {cfg_path.name}: {exc}")
    return changed_files


def _override_all_config_gui_host(gui_host: str) -> List[Path]:
    target = gui_host.strip()
    if not target:
        return []
    config_dir = _get_config_dir()
    if not config_dir.exists():
        return []
    changed_files: List[Path] = []
    candidates = sorted(config_dir.glob("*.xlsx")) + sorted(config_dir.glob("*.csv"))
    for cfg_path in candidates:
        try:
            if _overwrite_gui_settings_in_file(cfg_path, "", gui_host=target):
                changed_files.append(cfg_path)
        except Exception as exc:
            print(f"[WARN] Failed to update GUI HOST in {cfg_path.name}: {exc}")
    return changed_files


def _is_valid_ipv4(ip: str) -> bool:
    text = (ip or "").strip()
    if not text:
        return False
    try:
        socket.inet_aton(text)
    except Exception:
        return False
    return text.count(".") == 3


def _try_get_local_ipv4() -> Optional[str]:
    try:
        with socket.socket(socket.AF_INET, socket.SOCK_DGRAM) as sock:
            sock.connect(("8.8.8.8", 80))
            ip = sock.getsockname()[0]
        if _is_valid_ipv4(ip) and not ip.startswith("127."):
            return ip
    except Exception:
        pass
    try:
        ip = socket.gethostbyname(socket.gethostname())
        if _is_valid_ipv4(ip) and not ip.startswith("127."):
            return ip
    except Exception:
        pass
    return None


def _get_local_ipv4_for_gui() -> str:
    while True:
        ip = _try_get_local_ipv4()
        if ip:
            return ip

        # Pause and let user choose what to do when auto detection fails.
        try:
            root = tk.Tk()
            root.withdraw()
            root.attributes("-topmost", True)
            choice = messagebox.askyesnocancel(
                "HostIP Detect Failed",
                (
                    "Failed to get local IP.\n\n"
                    "Yes: Retry\n"
                    "No: Input HostIP manually\n"
                    "Cancel: Use default HostIP"
                ),
                parent=root,
            )
            root.destroy()
            if choice is True:
                continue
            if choice is False:
                manual = simpledialog.askstring(
                    "Manual HostIP",
                    "Input HostIP (IPv4):",
                )
                if _is_valid_ipv4(manual or ""):
                    return (manual or "").strip()
                messagebox.showwarning("Invalid HostIP", "Input is not a valid IPv4 address.")
                continue
            return GUI_HOST
        except Exception:
            pass

        print("[WARN] Failed to get local IP automatically.")
        print(f"[WARN] 1) Retry  2) Manual input  3) Use default ({GUI_HOST})")
        try:
            select = input("Select [3]: ").strip()
        except Exception:
            select = "3"
        if select == "1":
            continue
        if select == "2":
            try:
                manual = input("Input HostIP (IPv4): ").strip()
            except Exception:
                manual = ""
            if _is_valid_ipv4(manual):
                return manual
            print("[WARN] Invalid IPv4, please choose again.")
            continue
        return GUI_HOST


def _sync_gui_main_ini(exe_path: Path, host_ip: str, port: int = 7481) -> Optional[Path]:
    config_path = exe_path.parent / "Config" / "main.ini"
    if not config_path.exists():
        return None
    try:
        lines = config_path.read_text(encoding="utf-8", errors="ignore").splitlines()
    except Exception as exc:
        print(f"[WARN] Failed to read GUI main.ini: {exc}")
        return None

    target_map = {
        "TcpAutoCtrlServer": "1",
        "HostIP": str(host_ip).strip(),
        "Port": str(int(port)),
    }
    seen: set = set()
    out_lines: List[str] = []

    for line in lines:
        stripped = line.strip()
        if not stripped or stripped.startswith("#") or stripped.startswith(";") or "=" not in line:
            out_lines.append(line)
            continue
        key, _ = line.split("=", 1)
        key_clean = key.strip()
        if key_clean in target_map:
            out_lines.append(f"{key_clean}={target_map[key_clean]}")
            seen.add(key_clean)
        else:
            out_lines.append(line)

    for key in ("TcpAutoCtrlServer", "HostIP", "Port"):
        if key not in seen:
            out_lines.append(f"{key}={target_map[key]}")

    new_text = "\n".join(out_lines)
    old_text = "\n".join(lines)
    if new_text != old_text:
        try:
            config_path.write_text(new_text + "\n", encoding="utf-8")
            print(f"[INFO] Updated GUI main.ini: {config_path}")
        except Exception as exc:
            print(f"[WARN] Failed to write GUI main.ini: {exc}")
            return None
    return config_path


def _prompt_cal_power_scope() -> Tuple[Optional[float], Optional[float], Optional[float]]:
    try:
        root = tk.Tk()
        root.withdraw()
        root.attributes("-topmost", True)
        dialog = tk.Toplevel(root)
        dialog.title("Cal Power Scope")
        dialog.attributes("-topmost", True)
        dialog.resizable(False, False)

        enable_var = tk.BooleanVar(value=False)
        tk.Checkbutton(dialog, text="Enable Cal Power Scope", variable=enable_var).grid(
            row=0, column=0, columnspan=4, sticky="w", padx=12, pady=(12, 6)
        )

        tk.Label(dialog, text="Min").grid(row=1, column=0, padx=12, sticky="w")
        min_var = tk.StringVar(value="-1")
        min_entry = tk.Entry(dialog, textvariable=min_var, width=8)
        min_entry.grid(row=1, column=1, padx=6, pady=4)

        tk.Label(dialog, text="Max").grid(row=1, column=2, padx=6, sticky="w")
        max_var = tk.StringVar(value="0")
        max_entry = tk.Entry(dialog, textvariable=max_var, width=8)
        max_entry.grid(row=1, column=3, padx=6, pady=4)

        tk.Label(dialog, text="Step").grid(row=2, column=0, padx=12, sticky="w")
        step_var = tk.StringVar(value="1")
        step_entry = tk.Entry(dialog, textvariable=step_var, width=8)
        step_entry.grid(row=2, column=1, padx=6, pady=4)

        def _set_state(enabled: bool) -> None:
            state = "normal" if enabled else "disabled"
            min_entry.configure(state=state)
            max_entry.configure(state=state)
            step_entry.configure(state=state)

        _set_state(False)
        enable_var.trace_add("write", lambda *_: _set_state(enable_var.get()))

        result = {"value": None}

        def _parse_value(text: str) -> Optional[float]:
            raw = text.strip()
            if not raw:
                return None
            try:
                return float(raw)
            except Exception:
                return None

        def on_ok():
            if enable_var.get():
                result["value"] = (
                    _parse_value(min_var.get()),
                    _parse_value(max_var.get()),
                    _parse_value(step_var.get()),
                )
            else:
                result["value"] = (None, None, None)
            dialog.destroy()

        def on_cancel():
            result["value"] = (None, None, None)
            dialog.destroy()

        btn_frame = tk.Frame(dialog)
        btn_frame.grid(row=3, column=0, columnspan=4, pady=12)
        tk.Button(btn_frame, text="OK", width=8, command=on_ok).pack(side="left", padx=6)
        tk.Button(btn_frame, text="Cancel", width=8, command=on_cancel).pack(
            side="left", padx=6
        )

        dialog.protocol("WM_DELETE_WINDOW", on_cancel)
        dialog.grab_set()
        root.wait_window(dialog)
        root.destroy()
        return result["value"] or (None, None, None)
    except Exception:
        try:
            enable_text = input("Enable Cal Power Scope? (y/n): ").strip().lower()
            enable = enable_text in {"y", "yes"}
        except Exception:
            enable = False
        if not enable:
            return None, None, None
        try:
            scope_text = input("Cal Power Scope (min max, e.g. -1 0): ").strip()
            parts = scope_text.replace(",", " ").split()
            scope_min = float(parts[0]) if len(parts) > 0 else None
            scope_max = float(parts[1]) if len(parts) > 1 else None
        except Exception:
            scope_min = None
            scope_max = None
        try:
            step_text = input("Cal Power Step (e.g. 1): ").strip()
            scope_step = float(step_text) if step_text else None
        except Exception:
            scope_step = None
        return scope_min, scope_max, scope_step


def _read_table_rows(path: str) -> List[List[str]]:
    ext = Path(path).suffix.lower()
    if ext == ".xlsx":
        try:
            from openpyxl import load_workbook
        except Exception as exc:
            raise RuntimeError("openpyxl is required to read .xlsx files") from exc
        wb = load_workbook(path, data_only=True)
        ws = wb.active
        rows: List[List[str]] = []
        for row in ws.iter_rows(values_only=True):
            rows.append(["" if cell is None else str(cell) for cell in row])
        return rows
    with open(path, "r", newline="") as f:
        return list(csv.reader(f))


def _write_table_rows(path: str, rows: List[List[str]], template_path: Optional[str] = None) -> None:
    ext = Path(path).suffix.lower()
    if ext == ".xlsx":
        try:
            from openpyxl import Workbook
        except Exception as exc:
            raise RuntimeError("openpyxl is required to write .xlsx files") from exc
        wb = Workbook()
        ws = wb.active
        for row in rows:
            ws.append(list(row))
        wb.save(path)
        if template_path:
            _apply_xlsx_style(path, template_path)
        return
    with open(path, "w", newline="") as f:
        writer = csv.writer(f)
        writer.writerows(rows)


def _apply_xlsx_style(output_path: str, template_path: str) -> None:
    if not os.path.exists(template_path):
        return
    try:
        from openpyxl import load_workbook
    except Exception:
        return
    tpl_wb = None
    out_wb = None
    try:
        tpl_wb = load_workbook(template_path)
        tpl_ws = tpl_wb.active
        out_wb = load_workbook(output_path)
        out_ws = out_wb.active

        for col_letter, dim in tpl_ws.column_dimensions.items():
            if dim.width:
                out_ws.column_dimensions[col_letter].width = dim.width

        default_cell = tpl_ws.cell(1, 1)

        for row in out_ws.iter_rows():
            for cell in row:
                if cell.value is None:
                    continue
                tpl_cell = tpl_ws.cell(cell.row, cell.column)
                src = tpl_cell if tpl_cell.value is not None else default_cell
                cell.font = copy(src.font)
                cell.fill = copy(src.fill)
                cell.border = copy(src.border)
                cell.alignment = copy(src.alignment)
                cell.number_format = src.number_format

        out_wb.save(output_path)
    finally:
        if tpl_wb is not None:
            try:
                tpl_wb.close()
            except Exception:
                pass
        if out_wb is not None:
            try:
                out_wb.close()
            except Exception:
                pass


def _ensure_xlsx_support(paths: List[str]) -> bool:
    needs_xlsx = any(Path(p).suffix.lower() == ".xlsx" for p in paths)
    if not needs_xlsx:
        return True
    try:
        import openpyxl  # noqa: F401
    except Exception:
        print("[ERROR] openpyxl is required to read/write .xlsx files.")
        print("[ERROR] Install it with: pip install openpyxl")
        return False
    return True


def _load_cable_loss_table(path: str) -> List[Tuple[float, float, float]]:
    if not os.path.exists(path):
        return []
    pattern = re.compile(
        r"^CableLoss(?P<start>\d+(?:\.\d+)?)_(?P<end>\d+(?:\.\d+)?)G\s*=\s*(?P<loss>-?\d+(?:\.\d+)?)\s*$",
        re.IGNORECASE,
    )
    table: List[Tuple[float, float, float]] = []
    with open(path, "r", newline="") as f:
        for line in f:
            raw = line.strip()
            if not raw or raw.startswith("#"):
                continue
            match = pattern.match(raw)
            if not match:
                continue
            start_ghz = float(match.group("start"))
            end_ghz = float(match.group("end"))
            loss_db = float(match.group("loss"))
            table.append((start_ghz * 1e9, end_ghz * 1e9, loss_db))
    return table


def _lookup_cable_loss_db(
    freq_hz: float,
    table: List[Tuple[float, float, float]],
) -> Optional[float]:
    for start_hz, end_hz, loss_db in table:
        if start_hz <= freq_hz < end_hz:
            return loss_db
    for start_hz, end_hz, loss_db in table:
        if abs(freq_hz - end_hz) < 1e-6:
            return loss_db
    return None


def _load_global_csv_settings(
    path: str,
) -> Tuple[Optional[str], Optional[int], Optional[str], Optional[str]]:
    try:
        rows = _read_table_rows(path)
        if Path(path).suffix.lower() == ".xlsx":
            header = None
            values = None
            for row in rows:
                if not row or all(not str(c).strip() for c in row):
                    continue
                if header is None:
                    header = row
                    continue
                values = row
                break
            if header and values:
                host = None
                port = None
                connect_type = None
                gui_exe_path = None
                for i, key in enumerate(header):
                    key_str = str(key).strip()
                    if not key_str:
                        continue
                    key_clean = key_str.upper().replace("_", " ")
                    val = ""
                    if i < len(values):
                        val = str(values[i]).strip()
                    if not val:
                        continue
                    if key_clean == "GUI ADDRESS":
                        gui_exe_path = val
                    elif key_clean == "GUI HOST":
                        host = val
                    elif key_clean == "GUI PORT":
                        try:
                            port = int(float(val))
                        except ValueError:
                            pass
                    elif key_clean == "CONNECT TYPE":
                        connect_type = val.strip().upper()
                if host or port or connect_type or gui_exe_path:
                    return host, port, connect_type, gui_exe_path
        for row in rows:
            if not row or all(not str(c).strip() for c in row):
                continue
            host = None
            port = None
            connect_type = None
            for cell in row:
                cell_str = str(cell).strip()
                if ":" not in cell_str:
                    continue
                key, value = cell_str.split(":", 1)
                key_clean = key.strip().upper().replace("_", " ")
                val = value.strip()
                if not val:
                    continue
                if key_clean == "GUI HOST":
                    host = val
                elif key_clean == "GUI PORT":
                    try:
                        port = int(float(val))
                    except ValueError:
                        pass
                elif key_clean == "CONNECT TYPE":
                    connect_type = val.strip().upper()
            return host, port, connect_type, None
        return None, None, None, None
    except FileNotFoundError:
        return None, None, None, None


def _parse_harmonic_label(label: str) -> Optional[Tuple[float, str]]:
    raw = label.strip()
    if not raw:
        return None
    if "/" in raw:
        try:
            return float(Fraction(raw)), raw
        except Exception:
            return None
    low = raw.lower()
    for suffix in ("st", "nd", "rd", "th"):
        if low.endswith(suffix):
            raw = raw[:-len(suffix)]
            break
    try:
        return float(raw), label
    except ValueError:
        return None

def _build_tx_config_cmd(config: Dict[str, object]) -> str:
    return (
        "TX_CONFIG "
        f"CHAN {config['CHAN']} "
        f"BW {config['BW']} "
        f"OFFSET {config['OFFSET']} "
        f"MODE {config['MODE']} "
        f"OFDM_MODE {config['OFDM_MODE']} "
        f"RATE {config['RATE']} "
        f"CODING {config['CODING']} "
        f"DUTY_CYCLE {config['DUTY_CYCLE']} "
        f"PSDU_LEN {config['PSDU_LEN']}"
    )


def _is_bt_packet(packet_type: str) -> bool:
    upper = packet_type.strip().upper()
    return upper.startswith("BT_") and not upper.startswith("BT_BLE_")


def _build_bt_tx_config_cmd(config: Dict[str, object]) -> str:
    packet_type = _normalize_bt_packet_type(config.get("PACKET_TYPE", ""))
    if packet_type not in BT_PACKET_TYPES:
        raise ValueError(f"Invalid BT PACKET_TYPE: {packet_type!r}")
    chan = int(config.get("CHAN", 0))
    if _is_bt_packet(packet_type):
        if chan < 0 or chan > 78:
            raise ValueError(f"BT channel out of range for {packet_type}: {chan}")
    else:
        if chan < 0 or chan > 39:
            raise ValueError(f"BLE channel out of range for {packet_type}: {chan}")
    payload = int(config.get("PAYLOAD", 0))
    payload_len = int(config.get("PAYLOAD_LEN", 37))
    if payload < 0 or payload > 7:
        raise ValueError(f"PAYLOAD out of range: {payload}")
    if payload_len < 0 or payload_len > 255:
        raise ValueError(f"PAYLOAD_LEN out of range: {payload_len}")
    return (
        "BT_TX_CONFIG "
        f"PACKET_TYPE {packet_type} "
        f"CHAN {chan} "
        f"PAYLOAD {payload} "
        f"PAYLOAD_LEN {payload_len}"
    )


def _build_bt_rx_config_cmd(config: Dict[str, object]) -> str:
    packet_type = _normalize_bt_packet_type(config.get("PACKET_TYPE", ""))
    if packet_type not in BT_PACKET_TYPES:
        raise ValueError(f"Invalid BT PACKET_TYPE: {packet_type!r}")
    chan = int(config.get("CHAN", 0))
    if _is_bt_packet(packet_type):
        if chan < 0 or chan > 78:
            raise ValueError(f"BT channel out of range for {packet_type}: {chan}")
    else:
        if chan < 0 or chan > 39:
            raise ValueError(f"BLE channel out of range for {packet_type}: {chan}")
    return f"BT_RX_CONFIG PACKET_TYPE {packet_type} CHAN {chan}"


def measure_cpow_with_bt_power_calibration(
    inst: FsvSocket,
    gui,
    f0_hz: float,
    tolerance_db: float = 0.5,
    step_db: float = 0.25,
    max_iters: int = 20,
    control_tx: bool = True,
    desired_target: Optional[float] = None,
    cmd_delay: float = 0.0,
    loss_table: Optional[List[Tuple[float, float, float]]] = None,
    channel_bw_hz: float = 2e6,
) -> Tuple[float, float, float]:
    resp = gui.query("BT_POWER_GET")
    _sleep_cmd(cmd_delay)
    if desired_target is None:
        desired_target = _parse_power_value(resp)
    if desired_target is None:
        raise ValueError("Missing desired BT power target")
    current_target = desired_target
    gui.send(f"BT_POWER_TARGET POWER {current_target}")
    _sleep_cmd(cmd_delay)
    if control_tx:
        gui.send("BT_START_TX")
        _sleep_cmd(cmd_delay)
        if TX_START_STABLE_S > 0:
            time.sleep(TX_START_STABLE_S)
    try:
        cpow = measure_cpow_20m(inst, f0_hz, loss_table=loss_table, channel_bw_hz=channel_bw_hz)
        if abs(cpow - desired_target) > tolerance_db:
            current_target = desired_target + (desired_target - cpow)
            current_target = round(current_target / step_db) * step_db
            gui.send(f"BT_POWER_TARGET POWER {current_target}")
            _sleep_cmd(cmd_delay)
            cpow = measure_cpow_20m(inst, f0_hz, loss_table=loss_table, channel_bw_hz=channel_bw_hz)
            if abs(cpow - desired_target) > tolerance_db:
                for _ in range(max_iters):
                    if cpow < desired_target:
                        current_target += step_db
                    else:
                        current_target -= step_db
                    current_target = round(current_target / step_db) * step_db
                    gui.send(f"BT_POWER_TARGET POWER {current_target}")
                    _sleep_cmd(cmd_delay)
                    cpow = measure_cpow_20m(
                        inst, f0_hz, loss_table=loss_table, channel_bw_hz=channel_bw_hz
                    )
                    if abs(cpow - desired_target) <= tolerance_db:
                        break
        final_target = _parse_power_value(gui.query("BT_POWER_GET"))
        _sleep_cmd(cmd_delay)
        return desired_target, cpow, final_target
    finally:
        if control_tx:
            gui.send("BT_STOP_TX")
            _sleep_cmd(cmd_delay)

def _extract_harmonic_columns(header: List[str]) -> List[Tuple[float, str]]:
    known = {"CH", "FREQ", "PWR", "PWRTAR", "PWR_BT", "PWRTAR_BT"}
    results: List[Tuple[float, str]] = []
    for col in header:
        if not col:
            continue
        upper = col.strip().upper()
        if upper in known:
            continue
        if _normalize_key(col):
            continue
        parsed = _parse_harmonic_label(col)
        if parsed:
            results.append(parsed)
    return results

def _get_cell(row_map: Dict[str, str], key: str) -> str:
    for k, v in row_map.items():
        if k.strip().upper() == key.upper():
            return v
    return ""

def _set_cell(row_map: Dict[str, str], key: str, value: str) -> None:
    for k in row_map.keys():
        if k.strip().upper() == key.upper():
            row_map[k] = value
            return

def _build_cal_power_offsets(
    scope_min: Optional[float],
    scope_max: Optional[float],
    step: Optional[float],
) -> List[float]:
    if scope_min is None or scope_max is None or step is None or step <= 0:
        return [0.0]
    start = scope_min
    end = scope_max
    if start > end:
        start, end = end, start
    offsets: List[float] = []
    cur = start
    max_iters = 2000
    eps = 1e-9
    while cur <= end + eps and len(offsets) < max_iters:
        offsets.append(cur)
        cur += step
    if not offsets:
        return [0.0]
    return offsets


BANDEDGE_LIMIT_DBM = -46.0


def _extract_bandedge_columns(header: List[str]) -> List[Tuple[str, float]]:
    cols: List[Tuple[str, float]] = []
    for col in header:
        key = str(col or "").strip()
        if not key:
            continue
        upper = key.upper()
        if upper in {"CH", "FREQ", "PWR", "PWRTAR"}:
            continue
        if _normalize_key(key):
            continue
        try:
            freq_mhz = float(key)
        except Exception:
            continue
        if 1000.0 <= freq_mhz <= 7000.0:
            cols.append((key, freq_mhz * 1e6))
    return cols


def _pick_nearest_bandedge(f0_hz: float, edges: List[Tuple[str, float]]) -> Optional[Tuple[str, float, str]]:
    if not edges:
        return None
    best_label = ""
    best_edge_hz = 0.0
    best_dist = float("inf")
    for label, edge_hz in edges:
        dist = abs(edge_hz - f0_hz)
        if dist < best_dist:
            best_label, best_edge_hz, best_dist = label, edge_hz, dist
    if best_edge_hz <= 0:
        return None
    side = "LEFT" if best_edge_hz < f0_hz else "RIGHT"
    return best_label, best_edge_hz, side


def _tune_bandedge_max_power_target(
    inst: FsvSocket,
    gui,
    f0_hz: float,
    edge_hz: float,
    side: str,
    channel_bw_hz: float,
    loss_table: Optional[List[Tuple[float, float, float]]] = None,
    limit_dbm: float = BANDEDGE_LIMIT_DBM,
    step_db: float = 0.25,
    max_iters: int = 80,
) -> Tuple[float, float, float, float]:
    current_target = round(_parse_power_value(gui.power_get()) / step_db) * step_db
    _sleep_cmd(CMD_DELAY)
    # Stage 1 (coarse): assume 1 dB target change ~= 2 dB bandedge change.
    # Use 1 dB coarse steps until the edge power is within +/-2 dB around limit.
    coarse_target = current_target
    coarse_edge = None
    coarse_max_iters = min(max_iters, 20)
    for _ in range(coarse_max_iters):
        gui.power_target(coarse_target)
        _sleep_cmd(CMD_DELAY)
        edge_res = measure_bandedge_side_max(
            inst,
            _ensure_fsv_initialized,
            edge_hz=edge_hz,
            side="LEFT" if side.upper() == "LEFT" else "RIGHT",
            loss_table=loss_table,
        )
        coarse_edge = float(edge_res["power_dbm"])
        if abs(coarse_edge - limit_dbm) <= 2.0:
            break
        predicted = (limit_dbm - coarse_edge) / 2.0
        if predicted > 0:
            delta_db = max(1.0, round(predicted))
        else:
            delta_db = min(-1.0, round(predicted))
        coarse_target = coarse_target + float(delta_db)

    # Stage 2 (fine): 0.25 dB walk to find maximum passing target.
    current_target = round(coarse_target / step_db) * step_db
    best_target = None
    best_edge = None
    last_edge = coarse_edge
    last_target = current_target
    fine_max_iters = max(1, max_iters - coarse_max_iters)
    for _ in range(fine_max_iters):
        gui.power_target(current_target)
        _sleep_cmd(CMD_DELAY)
        edge_res = measure_bandedge_side_max(
            inst,
            _ensure_fsv_initialized,
            edge_hz=edge_hz,
            side="LEFT" if side.upper() == "LEFT" else "RIGHT",
            loss_table=loss_table,
        )
        edge_dbm = float(edge_res["power_dbm"])
        last_target = current_target
        last_edge = edge_dbm
        if edge_dbm <= limit_dbm:
            best_target = current_target
            best_edge = edge_dbm
            current_target = round((current_target + step_db) / step_db) * step_db
            continue
        if best_target is None:
            current_target = round((current_target - step_db) / step_db) * step_db
            continue
        break
    if best_target is None:
        final_target = float(last_target)
        final_edge = float(last_edge or 0.0)
    else:
        final_target = float(best_target)
        final_edge = float(best_edge or 0.0)
    gui.power_target(final_target)
    _sleep_cmd(CMD_DELAY)
    # Bandedge flow: solve target by edge limit first, then measure in-band power once.
    final_edge_res = measure_bandedge_side_max(
        inst,
        _ensure_fsv_initialized,
        edge_hz=edge_hz,
        side="LEFT" if side.upper() == "LEFT" else "RIGHT",
        loss_table=loss_table,
    )
    final_edge_dbm = float(final_edge_res["power_dbm"])
    final_edge_freq_hz = float(final_edge_res["freq_hz"])
    # Enforce "closest value below or equal to limit": if measured edge is above limit,
    # back off target by 0.25 dB and re-check until pass or iteration cap.
    if final_edge_dbm > limit_dbm:
        backoff_iters = 40
        for _ in range(backoff_iters):
            final_target = round((final_target - step_db) / step_db) * step_db
            gui.power_target(final_target)
            _sleep_cmd(CMD_DELAY)
            chk = measure_bandedge_side_max(
                inst,
                _ensure_fsv_initialized,
                edge_hz=edge_hz,
                side="LEFT" if side.upper() == "LEFT" else "RIGHT",
                loss_table=loss_table,
            )
            final_edge_dbm = float(chk["power_dbm"])
            final_edge_freq_hz = float(chk["freq_hz"])
            if final_edge_dbm <= limit_dbm:
                break
    final_cpow = measure_cpow_20m(inst, f0_hz, loss_table=loss_table, channel_bw_hz=channel_bw_hz)
    return (
        float(final_cpow),
        final_target,
        final_edge_dbm,
        final_edge_freq_hz,
    )


def run_bandedge_test(
    input_path: str,
    output_path: str,
    inst: FsvSocket,
    gui,
    default_connect_type: Optional[str] = None,
    default_firmware_type: Optional[str] = None,
    cable_loss_table: Optional[List[Tuple[float, float, float]]] = None,
    band_24: bool = True,
    band_5: bool = True,
    profile: str = "SINGLE_BAND",
    flow_mode: str = "WIFI",
) -> None:
    rows_out: List[List[str]] = []
    base_defaults = dict(DEFAULT_TX_CONFIG)
    if default_connect_type:
        base_defaults["CONNECT_TYPE"] = default_connect_type
    base_defaults["FIRMWARE_TYPE"] = "WIFI" if flow_mode.upper() != "BT" else "BLE"
    if default_firmware_type in FIRMWARE_TYPES:
        base_defaults["FIRMWARE_TYPE"] = default_firmware_type
    defaults = dict(base_defaults)

    header: Optional[List[str]] = None
    bandedge_cols: List[Tuple[str, float]] = []
    pending_config_header: Optional[List[str]] = None
    pending_config_header_row: Optional[List[str]] = None
    pending_config_value_row: Optional[List[str]] = None
    pending_header_row: Optional[List[str]] = None
    current_connect_type: Optional[str] = None
    current_antenna: Optional[str] = None
    current_cert_mode: Optional[str] = None
    current_gpio20: Optional[str] = None
    current_gpio21: Optional[str] = None
    current_firmware_type: Optional[str] = None
    loss_table = cable_loss_table or []
    profile_key = (profile or "SINGLE_BAND").strip().upper()

    def _flush_pending_block_headers() -> None:
        nonlocal pending_config_header_row, pending_config_value_row, pending_header_row
        if pending_config_header_row:
            rows_out.append(pending_config_header_row)
            pending_config_header_row = None
        if pending_config_value_row:
            rows_out.append(pending_config_value_row)
            pending_config_value_row = None
        if pending_header_row:
            rows_out.append(pending_header_row)
            pending_header_row = None

    def _process_row(row_map: Dict[str, str], config: Dict[str, object], f0_hz: float) -> None:
        nonlocal current_connect_type, current_antenna, current_cert_mode
        nonlocal current_gpio20, current_gpio21, current_firmware_type
        if not header:
            return
        chosen = _pick_nearest_bandedge(f0_hz, bandedge_cols)
        channel_bw_hz = _parse_wifi_channel_bw_hz(config.get("BW"), default_hz=20e6)
        if chosen is None:
            _flush_pending_block_headers()
            rows_out.append([row_map.get(col, "") for col in header])
            return
        edge_label, edge_hz, side = chosen
        cpow = 0.0
        pwr_tar = None
        edge_dbm = None
        marker_freq_hz = None
        started_tx = False
        if gui:
            firmware_type = _normalize_firmware_type(config.get("FIRMWARE_TYPE", ""), default="WIFI")
            if firmware_type != current_firmware_type:
                gui.send(f"SELECT_FIRMWARE_TYPE TYPE {firmware_type}")
                _sleep_cmd(FW_SWITCH_SETTLE_S)
                current_firmware_type = firmware_type
                current_connect_type = None
            connect_type = _normalize_connect_type(str(config.get("CONNECT_TYPE", "")))
            if connect_type and connect_type != current_connect_type:
                gui.send(f"CONNECT TYPE {connect_type}")
                _sleep_cmd(max(CMD_DELAY, CONNECT_SETTLE_S))
                current_connect_type = connect_type
            antenna = str(config.get("ANTENNA", "")).strip().upper()
            cert_mode = str(config.get("CERTIFICATION_MODE", "")).strip().upper()
            if antenna and antenna != current_antenna:
                gui.send(f"ANTENNA ANT {antenna}")
                _sleep_cmd(CMD_DELAY)
                current_antenna = antenna
            if cert_mode and cert_mode != current_cert_mode:
                gui.send(f"CERTIFICATION MODE {cert_mode}")
                _sleep_cmd(CMD_DELAY)
                current_cert_mode = cert_mode
            if profile_key == "DULE_ANTENNA" and antenna in {"ANT1", "ANT2"}:
                gpio20_level = "HIGH" if antenna == "ANT1" else "LOW"
                gpio21_level = "LOW" if antenna == "ANT1" else "HIGH"
            else:
                gpio20_level = _normalize_gpio_level(config.get("GPIO20"))
                gpio21_level = _normalize_gpio_level(config.get("GPIO21"))
            if gpio20_level and gpio20_level != current_gpio20:
                gui.send(f"GPIO_OUTPUT GPIO 20 LEVEL {gpio20_level}")
                _sleep_cmd(CMD_DELAY)
                current_gpio20 = gpio20_level
            if gpio21_level and gpio21_level != current_gpio21:
                gui.send(f"GPIO_OUTPUT GPIO 21 LEVEL {gpio21_level}")
                _sleep_cmd(CMD_DELAY)
                current_gpio21 = gpio21_level

            gui.send(_build_tx_config_cmd(config))
            _sleep_cmd(CMD_DELAY)
            gui.start_tx()
            _sleep_cmd(CMD_DELAY)
            started_tx = True
            cpow, pwr_tar, edge_dbm, marker_freq_hz = _tune_bandedge_max_power_target(
                inst,
                gui,
                f0_hz=f0_hz,
                edge_hz=edge_hz,
                side=side,
                channel_bw_hz=channel_bw_hz,
                loss_table=loss_table,
            )
        else:
            cpow = measure_cpow_20m(
                inst,
                f0_hz,
                loss_table=loss_table,
                channel_bw_hz=channel_bw_hz,
            )
            edge_res = measure_bandedge_side_max(
                inst,
                _ensure_fsv_initialized,
                edge_hz=edge_hz,
                side="LEFT" if side == "LEFT" else "RIGHT",
                loss_table=loss_table,
            )
            edge_dbm = float(edge_res["power_dbm"])
            marker_freq_hz = float(edge_res["freq_hz"])

        row_map["Pwr"] = f"{cpow:.1f}"
        if pwr_tar is not None:
            row_map["PwrTar"] = f"{pwr_tar:.1f}"
        if edge_dbm is not None:
            row_map[edge_label] = f"{edge_dbm:.1f}"
        if marker_freq_hz is not None:
            row_map["MarkerFreq"] = f"{marker_freq_hz/1e6:.1f}"
        _flush_pending_block_headers()
        rows_out.append([row_map.get(col, "") for col in header])
        if gui and started_tx:
            gui.stop_tx()
            _sleep_cmd(CMD_DELAY)

    rows = _read_table_rows(input_path)
    gui_rows: set = set()
    for idx, row in enumerate(rows):
        if not row or all(not str(c).strip() for c in row):
            continue
        upper = [str(c).strip().upper() for c in row]
        if "GUI_ADDRESS" in upper or "GUI HOST" in upper:
            gui_header_idx = idx
            gui_value_idx = None
            for j in range(idx + 1, len(rows)):
                r2 = rows[j]
                if not r2 or all(not str(c).strip() for c in r2):
                    continue
                gui_value_idx = j
                break
            if gui_value_idx is not None:
                gui_rows.add(gui_header_idx)
                gui_rows.add(gui_value_idx)
                rows_out.append(rows[gui_header_idx])
                rows_out.append(rows[gui_value_idx])
            break

    for idx, row in enumerate(rows):
        if idx in gui_rows:
            continue
        if not row or all(not str(c).strip() for c in row):
            rows_out.append(row)
            continue

        if pending_config_header is not None:
            defaults = dict(base_defaults)
            _apply_config_header_row(pending_config_header, row, defaults)
            header = None
            bandedge_cols = []
            pending_config_header = None
            pending_config_value_row = row
            continue

        if any(":" in c for c in row):
            defaults = dict(base_defaults)
            for cell in row:
                _parse_config_cell(cell, defaults)
            header = None
            bandedge_cols = []
            pending_config_header_row = None
            pending_config_value_row = None
            pending_header_row = None
            rows_out.append(row)
            continue

        if _is_config_header_row(row):
            # New block begins; drop previous pending block headers if it had no data rows.
            pending_config_header_row = None
            pending_config_value_row = None
            pending_header_row = None
            pending_config_header = row
            pending_config_header_row = row
            continue

        if row[0].strip().upper() == "CH":
            header = [c.strip() for c in row]
            if "MarkerFreq" not in header:
                header.append("MarkerFreq")
            marker_idx = header.index("MarkerFreq")
            target_idx = max(0, marker_idx - 2)
            if target_idx != marker_idx:
                marker_col = header.pop(marker_idx)
                header.insert(target_idx, marker_col)
            bandedge_cols = _extract_bandedge_columns(header)
            pending_header_row = list(header)
            continue

        if not header:
            rows_out.append(row)
            continue

        row_map = {header[i]: row[i].strip() if i < len(row) else "" for i in range(len(header))}
        if not _get_cell(row_map, "CH"):
            rows_out.append(row)
            continue

        config = dict(defaults)
        for col, val in row_map.items():
            if not val:
                continue
            norm = _normalize_key(col)
            if not norm:
                continue
            if norm == "CHAN":
                chan = _parse_chan(val)
                if chan is not None:
                    config[norm] = chan
                continue
            if norm in {"OFFSET"}:
                try:
                    config[norm] = float(val)
                except Exception:
                    pass
                continue
            if norm in {"DUTY_CYCLE", "PSDU_LEN", "PAYLOAD", "PAYLOAD_LEN"}:
                try:
                    config[norm] = int(float(val))
                except Exception:
                    pass
                continue
            if norm in {"FIRMWARE_TYPE", "TEST_MODE", "PACKET_TYPE"}:
                if norm == "FIRMWARE_TYPE":
                    config[norm] = _normalize_firmware_type(val, default="WIFI")
                elif norm == "PACKET_TYPE":
                    config[norm] = _normalize_bt_packet_type(val)
                else:
                    config[norm] = str(val).strip().upper()
                continue
            config[norm] = val

        antenna_cfg = str(config.get("ANTENNA", "")).strip().upper()
        if profile_key != "DULE_ANTENNA" and antenna_cfg == "ANT2":
            # SINGLE_BAND / DULE_BAND do not support ANT2 path in bandedge run.
            continue

        freq_cell = _get_cell(row_map, "Freq")
        if not freq_cell:
            rows_out.append(row)
            continue
        try:
            f0_hz = float(freq_cell) * 1e6
        except Exception:
            rows_out.append(row)
            continue
        if f0_hz < 3e9 and not band_24:
            continue
        if f0_hz >= 4e9 and not band_5:
            continue
        _process_row(row_map, config, f0_hz)

    _write_table_rows(output_path, rows_out, template_path=input_path)

def run_csv_test(
    input_path: str,
    output_path: str,
    inst: FsvSocket,
    gui,
    default_connect_type: Optional[str] = None,
    default_firmware_type: Optional[str] = None,
    cable_loss_table: Optional[List[Tuple[float, float, float]]] = None,
    cal_scope_min: Optional[float] = None,
    cal_scope_max: Optional[float] = None,
    cal_scope_step: Optional[float] = None,
    band_24: bool = True,
    band_5: bool = True,
    test_harmonic: bool = True,
    test_bandedge: bool = True,
    flow_mode: str = "WIFI",
) -> None:
    rows_out: List[List[str]] = []
    base_defaults = dict(DEFAULT_TX_CONFIG)
    if default_connect_type:
        base_defaults["CONNECT_TYPE"] = default_connect_type
    if default_firmware_type in FIRMWARE_TYPES:
        base_defaults["FIRMWARE_TYPE"] = default_firmware_type
    elif flow_mode.upper() == "BT":
        base_defaults["FIRMWARE_TYPE"] = "BLE"
    else:
        base_defaults["FIRMWARE_TYPE"] = "WIFI"
    defaults = dict(base_defaults)
    header: Optional[List[str]] = None
    harmonics: List[Tuple[float, str]] = []
    pending_config_header: Optional[List[str]] = None
    pending_config_header_row: Optional[List[str]] = None
    pending_config_value_row: Optional[List[str]] = None
    pending_header_row: Optional[List[str]] = None
    current_connect_type: Optional[str] = None
    current_antenna: Optional[str] = None
    current_cert_mode: Optional[str] = None
    current_gpio20: Optional[str] = None
    current_gpio21: Optional[str] = None
    current_firmware_type: Optional[str] = None
    last_calibration: Dict[Tuple[float, float], Tuple[float, float]] = {}
    loss_table = cable_loss_table or []
    use_scope = (
        flow_mode.upper() != "BT"
        and test_harmonic
        and
        cal_scope_min is not None
        and cal_scope_max is not None
        and cal_scope_step is not None
        and cal_scope_step > 0
        and gui is not None
    )
    run_harmonic_test = (flow_mode.upper() == "BT") or test_harmonic
    use_wifi_bandedge = (flow_mode.upper() != "BT") and test_bandedge
    offsets = _build_cal_power_offsets(cal_scope_min, cal_scope_max, cal_scope_step) if use_scope else [0.0]
    block_rows: List[Dict[str, object]] = []

    def _process_row(entry: Dict[str, object], offset: float, block_base_cal: Optional[float]) -> bool:
        nonlocal current_connect_type, current_antenna, current_cert_mode
        nonlocal current_gpio20, current_gpio21, current_firmware_type, last_calibration

        row_map = dict(entry["row_map"])
        config = entry["config"]
        f0_hz = entry["f0_hz"]
        channel_bw_hz = _parse_wifi_channel_bw_hz(config.get("BW"), default_hz=20e6)
        started_tx = False
        desired_target = None
        power_tar_final = None
        if gui:
            if (default_firmware_type or "").strip().upper() == "WIFI_AND_BLE":
                default_fw = "WIFI_AND_BLE"
            else:
                default_fw = "BLE" if flow_mode.upper() == "BT" else "WIFI"
            row_fw = _normalize_firmware_type(config.get("FIRMWARE_TYPE", ""), default=default_fw)
            if flow_mode.upper() == "BT":
                firmware_type = "WIFI_AND_BLE" if row_fw == "WIFI_AND_BLE" else "BLE"
            else:
                firmware_type = row_fw
            if firmware_type != current_firmware_type:
                # Some GUI builds auto-reconnect after DISCONNECT, so switch first.
                gui.send(f"SELECT_FIRMWARE_TYPE TYPE {firmware_type}")
                _sleep_cmd(FW_SWITCH_SETTLE_S)
                current_firmware_type = firmware_type
                current_connect_type = None
            connect_type = _normalize_connect_type(str(config.get("CONNECT_TYPE", "")))
            if connect_type and connect_type != current_connect_type:
                gui.send(f"CONNECT TYPE {connect_type}")
                _sleep_cmd(max(CMD_DELAY, CONNECT_SETTLE_S))
                current_connect_type = connect_type
            is_bt_path = flow_mode.upper() == "BT"

            antenna = str(config.get("ANTENNA", "")).strip().upper()
            cert_mode = str(config.get("CERTIFICATION_MODE", "")).strip().upper()
            if not is_bt_path:
                if antenna:
                    if antenna not in {"ANT1", "ANT2", "ALL"}:
                        raise ValueError(f"Invalid ANTENNA value: {antenna!r}")
                    if antenna != current_antenna:
                        gui.send(f"ANTENNA ANT {antenna}")
                        _sleep_cmd(CMD_DELAY)
                        current_antenna = antenna
                if cert_mode:
                    if cert_mode not in {"NORMAL", "CE", "FCC", "SRRC", "SRRC_2"}:
                        raise ValueError(f"Invalid CERTIFICATION MODE value: {cert_mode!r}")
                    if cert_mode != current_cert_mode:
                        gui.send(f"CERTIFICATION MODE {cert_mode}")
                        _sleep_cmd(CMD_DELAY)
                        if cert_mode == "FCC":
                            _sleep_cmd(0.5)
                        current_cert_mode = cert_mode
                gpio20_level = _normalize_gpio_level(config.get("GPIO20"))
                if gpio20_level and gpio20_level != current_gpio20:
                    gui.send(f"GPIO_OUTPUT GPIO 20 LEVEL {gpio20_level}")
                    _sleep_cmd(CMD_DELAY)
                    current_gpio20 = gpio20_level
                gpio21_level = _normalize_gpio_level(config.get("GPIO21"))
                if gpio21_level and gpio21_level != current_gpio21:
                    gui.send(f"GPIO_OUTPUT GPIO 21 LEVEL {gpio21_level}")
                    _sleep_cmd(CMD_DELAY)
                    current_gpio21 = gpio21_level

            if is_bt_path:
                gui.send(_build_bt_tx_config_cmd(config))
                _sleep_cmd(max(CMD_DELAY, BT_TXCFG_TO_START_DELAY_S))
                gui.send("BT_START_TX")
                _sleep_cmd(CMD_DELAY)
                started_tx = True
            else:
                gui.send(_build_tx_config_cmd(config))
                _sleep_cmd(CMD_DELAY)
                gui.start_tx()
                _sleep_cmd(CMD_DELAY)
                started_tx = True
            if SIMPLE_GUI_FLOW:
                print("[INFO] SIMPLE_GUI_FLOW enabled: only TX_CONFIG -> START_TX, then stop.")
                return True
            if STOP_AFTER_START_TX:
                print("[INFO] STOP_AFTER_START_TX enabled: press Enter to continue (TX stays on).")
                try:
                    input()
                except KeyboardInterrupt:
                    print()
                return True
            skip_cal = False if is_bt_path else _should_skip_calibration(
                antenna, config.get("GPIO20"), config.get("GPIO21")
            )

            if use_scope and block_base_cal is not None:
                desired_target = block_base_cal + offset
            else:
                cal_pwr_cell = _get_cell(row_map, "Cal Pwr") or _get_cell(row_map, "Cal Power")
                if cal_pwr_cell:
                    try:
                        desired_target = float(cal_pwr_cell)
                    except Exception:
                        desired_target = None
                if desired_target is None:
                    ch_cell = _get_cell(row_map, "CH")
                    raise ValueError(f"Missing or invalid Cal Pwr for CH={ch_cell!r}")
                if use_scope:
                    desired_target = desired_target + offset

            if use_scope:
                _set_cell(row_map, "Cal Pwr", f"{desired_target:.1f}")
                _set_cell(row_map, "Cal Power", f"{desired_target:.1f}")

            cache_key = (f0_hz, desired_target)
            cached = last_calibration.get(cache_key)
            if skip_cal and cached:
                cached_pwr, cached_pwr_tar = cached
                if is_bt_path:
                    gui.send(f"BT_POWER_TARGET POWER {cached_pwr_tar}")
                    _sleep_cmd(CMD_DELAY)
                else:
                    gui.power_target(cached_pwr_tar)
                    _sleep_cmd(CMD_DELAY)
                cpow = cached_pwr
                power_tar_final = cached_pwr_tar
            else:
                if skip_cal and not cached and not use_scope:
                    print(
                        f"[WARN] Skip calibration requested but no prior data for {f0_hz/1e6:.0f} MHz; "
                        "falling back to calibration."
                    )
                if is_bt_path:
                    power_target, cpow, power_tar_final = measure_cpow_with_bt_power_calibration(
                        inst,
                        gui,
                        f0_hz,
                        control_tx=False,
                        desired_target=desired_target,
                        cmd_delay=CMD_DELAY,
                        loss_table=loss_table,
                    )
                else:
                    power_target, cpow, power_tar_final = measure_cpow_with_power_calibration(
                        inst,
                        gui,
                        f0_hz,
                        control_tx=False,
                        desired_target=desired_target,
                        cmd_delay=CMD_DELAY,
                        loss_table=loss_table,
                        channel_bw_hz=channel_bw_hz,
                    )
                last_calibration[cache_key] = (cpow, power_tar_final)
        else:
            cpow = measure_cpow_20m(
                inst,
                f0_hz,
                loss_table=loss_table,
                channel_bw_hz=channel_bw_hz,
            )

        row_map["Pwr"] = f"{cpow:.1f}"
        if flow_mode.upper() == "BT":
            row_map["Pwr_BT"] = f"{cpow:.1f}"
        if power_tar_final is not None:
            row_map["PwrTar"] = f"{power_tar_final:.1f}"
            if flow_mode.upper() == "BT":
                row_map["PwrTar_BT"] = f"{power_tar_final:.1f}"

        if STOP_AFTER_CALIBRATION:
            pwr_tar_text = f"{power_tar_final:.1f}" if power_tar_final is not None else "N/A"
            print("[INFO] STOP_AFTER_CALIBRATION enabled: current TX config and targets.")
            print(f"[INFO] TX_CONFIG: {config}")
            print(f"[INFO] Cal Pwr={desired_target} Pwr={cpow:.1f} PwrTar={pwr_tar_text}")
            return True

        if run_harmonic_test and harmonics:
            inst.query("SENS:FREQ:SPAN?")
            orders = [h[0] for h in harmonics]
            order_strs = [h[1] for h in harmonics]
            harm_results = measure_all_harmonics(
                inst,
                f0_hz,
                MAX_FREQ_HZ,
                [25, 20, 15],
                orders=orders,
                orders_str=order_strs,
                loss_table=loss_table,
            )
            inst.query("SENS:FREQ:SPAN?")
            for item in harm_results:
                key = item.get("order_str", str(item["order"]))
                row_map[key] = f"{item['best']['power']:.1f}"
                if flow_mode.upper() == "BT":
                    row_map[f"{key}_BT"] = f"{item['best']['power']:.1f}"

        if use_wifi_bandedge:
            bandedge_results = measure_wifi_bandedges(
                inst,
                f0_hz,
                _ensure_fsv_initialized,
                loss_table=loss_table,
            )
            for col in WIFI_BANDEDGE_COLUMNS:
                if col in bandedge_results:
                    row_map[col] = f"{bandedge_results[col]:.1f}"

        out_row = [row_map.get(col, "") for col in header]
        rows_out.append(out_row)
        if gui and started_tx:
            if flow_mode.upper() == "BT":
                gui.send("BT_STOP_TX")
                _sleep_cmd(CMD_DELAY)
            else:
                gui.stop_tx()
                _sleep_cmd(CMD_DELAY)
        return False

    def _process_block() -> bool:
        nonlocal pending_header_row
        nonlocal pending_config_header_row, pending_config_value_row
        if not block_rows:
            return False
        if pending_config_header_row:
            rows_out.append(pending_config_header_row)
        if pending_config_value_row:
            rows_out.append(pending_config_value_row)
        if pending_header_row:
            rows_out.append(pending_header_row)
        pending_config_header_row = None
        pending_config_value_row = None
        pending_header_row = None
        block_base_cal = None
        if use_scope:
            first_row_map = block_rows[0]["row_map"]
            cal_pwr_cell = _get_cell(first_row_map, "Cal Pwr") or _get_cell(first_row_map, "Cal Power")
            if cal_pwr_cell:
                try:
                    block_base_cal = float(cal_pwr_cell)
                except Exception:
                    block_base_cal = None
            if block_base_cal is None:
                raise ValueError("Missing or invalid Cal Pwr for scope block")
        for offset in offsets:
            for entry in block_rows:
                if _process_row(entry, offset, block_base_cal):
                    return True
        block_rows.clear()
        return False

    rows = _read_table_rows(input_path)
    gui_rows: set = set()
    for idx, row in enumerate(rows):
        if not row or all(not str(c).strip() for c in row):
            continue
        upper = [str(c).strip().upper() for c in row]
        if "GUI_ADDRESS" in upper or "GUI HOST" in upper:
            gui_header_idx = idx
            gui_value_idx = None
            for j in range(idx + 1, len(rows)):
                r2 = rows[j]
                if not r2 or all(not str(c).strip() for c in r2):
                    continue
                gui_value_idx = j
                break
            if gui_value_idx is not None:
                gui_rows.add(gui_header_idx)
                gui_rows.add(gui_value_idx)
                rows_out.append(rows[gui_header_idx])
                rows_out.append(rows[gui_value_idx])
            break

    for idx, row in enumerate(rows):
        if idx in gui_rows:
            continue
        if not row or all(not c.strip() for c in row):
            if _process_block():
                return
            rows_out.append(row)
            continue

        if pending_config_header is not None:
            if _process_block():
                return
            defaults = dict(base_defaults)
            _apply_config_header_row(pending_config_header, row, defaults)
            header = None
            harmonics = []
            pending_config_header = None
            pending_config_value_row = row
            continue

        if any(":" in c for c in row):
            if _process_block():
                return
            defaults = dict(base_defaults)
            for cell in row:
                _parse_config_cell(cell, defaults)
            header = None
            harmonics = []
            pending_config_value_row = row
            continue

        if _is_config_header_row(row):
            if _process_block():
                return
            pending_config_header = row
            pending_config_header_row = row
            continue

        if row[0].strip().upper() == "CH":
            if _process_block():
                return
            header = [c.strip() for c in row]
            if use_wifi_bandedge:
                for col in WIFI_BANDEDGE_COLUMNS:
                    if col not in header:
                        header.append(col)
            harmonics = _extract_harmonic_columns(header)
            pending_header_row = list(header)
            continue

        if not header:
            if _process_block():
                return
            rows_out.append(row)
            continue

        row_map = {header[i]: row[i].strip() if i < len(row) else "" for i in range(len(header))}
        if not _get_cell(row_map, "CH"):
            if _process_block():
                return
            rows_out.append(row)
            continue

        config = dict(defaults)
        for col, val in row_map.items():
            if not val:
                continue
            norm = _normalize_key(col)
            if not norm:
                continue
            if norm == "CHAN":
                chan = _parse_chan(val)
                if chan is not None:
                    config[norm] = chan
                continue
            if norm == "OFFSET":
                try:
                    config[norm] = float(val)
                except ValueError:
                    pass
                continue
            if norm in {"DUTY_CYCLE", "PSDU_LEN", "PAYLOAD", "PAYLOAD_LEN"}:
                try:
                    config[norm] = int(float(val))
                except ValueError:
                    pass
                continue
            if norm in {"FIRMWARE_TYPE", "TEST_MODE", "PACKET_TYPE"}:
                if norm == "FIRMWARE_TYPE":
                    config[norm] = _normalize_firmware_type(val, default="WIFI")
                elif norm == "PACKET_TYPE":
                    config[norm] = _normalize_bt_packet_type(val)
                else:
                    config[norm] = str(val).strip().upper()
                continue
            config[norm] = val

        freq_cell = _get_cell(row_map, "Freq")
        if not freq_cell:
            if _process_block():
                return
            rows_out.append(row)
            continue
        try:
            f0_hz = float(freq_cell) * 1e6
        except ValueError:
            if _process_block():
                return
            rows_out.append(row)
            continue

        if flow_mode.upper() != "BT":
            if f0_hz < 3e9:
                if not band_24:
                    continue
            elif f0_hz >= 4e9:
                if not band_5:
                    continue

        block_rows.append(
            {
                "row_map": row_map,
                "config": config,
                "f0_hz": f0_hz,
            }
        )

    if _process_block():
        return

    _write_table_rows(output_path, rows_out, template_path=input_path)
def measure_one_harmonic(
    inst: FsvSocket,
    f0_hz: float,
    order: float,
    att_list: List[int],
    span_hz: float = 100e6,
    ref_level_dbm: float = 10.0,
    order_str: Optional[str] = None,
    loss_table: Optional[List[Tuple[float, float, float]]] = None,
) -> Dict[str, Any]:
    """
    测某一个谐波 n*f0：
      - 多个 ATT 尝试
      - 每个 ATT：扫一枪 + Marker MAX
      - 返回每个 ATT 结果 + 选出的最大值
    """
    f_center = f0_hz * order
    display_order = order_str if order_str is not None else str(order)
    print(f"\n===== Step2: {display_order} 次谐波，中心 ≈ {f_center/1e9:.6f} GHz =====")

    # 切回普通扫频模式（离开 CPOW / ACP 等测量 app）
    inst.send_cmd("SENS:FREQ:MODE SWE")
    inst.send_cmd("INIT:CONT OFF")

    inst.send_cmd(f"SENS:FREQ:CENT {f_center:.0f}")
    inst.send_cmd(f"SENS:FREQ:SPAN {int(span_hz)}")

    # RBW / VBW = 1MHz（重申一遍也没问题）
    inst.send_cmd("SENS:BAND:AUTO OFF")
    inst.send_cmd("SENS:BAND 1MHz")
    inst.send_cmd("SENS:BAND:VID:AUTO OFF")
    inst.send_cmd("SENS:BAND:VID 1MHz")

    # Detector：RMS（可按需要改成 POS）
    inst.send_cmd("SENS:WIND:DET1:FUNC RMS")

    inst.send_cmd(f"DISP:WIND:TRAC:Y:SCAL:RLEV {ref_level_dbm}")
    if loss_table:
        loss_db = _lookup_cable_loss_db(f_center, loss_table)
        if loss_db is not None:
            inst.send_cmd(f"DISP:WIND:TRAC:Y:SCAL:RLEV:OFFS {loss_db}")
    inst.send_cmd("INP:ATT:AUTO OFF")

    inst.send_cmd("CALC:MARK1:STAT ON")

    per_att_results: List[Dict[str, float]] = []

    for att in att_list:
        print(f"[ATT LOOP] ATT = {att} dB")
        inst.send_cmd(f"INP:ATT {att}")

        inst.send_cmd("INIT:IMM")
        inst.send_cmd("*WAI")

        inst.send_cmd("CALC:MARK1:MAX")

        f_peak = inst.query_float("CALC:MARK1:X?")
        p_peak = inst.query_float("CALC:MARK1:Y?")

        inst.check_error(f"harmonic {order}, ATT={att}")

        per_att_results.append({
            "atten": att,
            "freq": f_peak,
            "power": p_peak,
        })

        print(f"  → 峰值: {f_peak/1e9:.6f} GHz, {p_peak:.2f} dBm")

    # 选功率最小的那一组
    best = min(per_att_results, key=lambda r: r["power"])

    display_order = order_str if order_str is not None else str(order)
    print(
        f"[BEST] {display_order} 次谐波: "
        f"ATT={best['atten']} dB, "
        f"f_peak={best['freq']/1e9:.6f} GHz, " 
        f"P_peak={best['power']:.2f} dBm"
    )

    return {
        "order": order,
        "order_str": order_str if order_str is not None else str(order),
        "center": f_center,
        "per_atten": per_att_results,
        "best": best,
    }


def measure_all_harmonics(
    inst: FsvSocket,
    f0_hz: float,
    max_freq_hz: float,
    att_list: List[int],
    orders: Optional[List[float]] = None,
    orders_str: Optional[List[str]] = None,
    loss_table: Optional[List[Tuple[float, float, float]]] = None,
) -> List[Dict[str, Any]]:
    """
    如果指定 orders（如 [2,3,4.5]），只测这些阶数（不限制频率上限）。
    否则默认从 2 次开始依次测到 n*f0 > max_freq_hz 为止。
    """
    results: List[Dict[str, Any]] = []
    if orders:
        # 手动输入谐波时，不限制频率上限
        for i, order in enumerate(orders):
            order_str = orders_str[i] if orders_str and i < len(orders_str) else None
            res = measure_one_harmonic(
                inst,
                f0_hz,
                order,
                att_list,
                order_str=order_str,
                loss_table=loss_table,
            )
            results.append(res)
    else:
        # 自动模式时，限制到 max_freq_hz
        order = 2
        while True:
            f_harm = f0_hz * order
            if f_harm > max_freq_hz:
                break
            res = measure_one_harmonic(
                inst,
                f0_hz,
                order,
                att_list,
                loss_table=loss_table,
            )
            results.append(res)
            order += 1
    return results



def export_simple_csv(
    filename: str,
    f0_hz: float,
    cpow: float,
    harm_results: List[Dict[str, Any]]
):
    """
    导出一个很简单的 CSV（Excel 可直接打开）：

    表头格式：Freq_GHz, CPOW_dBm, H2_dBm, H3_dBm, ...
    数据行：本次测试对应的数值，频率保留 3 位小数，功率保留 1 位小数。

    表头逻辑：
      - 如果文件不存在：写表头 + 数据
      - 如果文件存在：只和“最近一次写入的表头”（最后一行中是表头的那行）比较谐波列：
          * 若谐波列相同 → 只追加数据行，不写新表头
          * 若谐波列不同 → 先写新表头，再写数据行
    """

    order_list = [item["order"] for item in harm_results]
    order_str_list = [item.get("order_str", str(item["order"])) for item in harm_results]
    headers = ["Freq_GHz", "CPOW_dBm"] + [
        f"H{order_str}_dBm" for order_str in order_str_list
    ]

    order2power = {
        item["order"]: round(item["best"]["power"], 1)
        for item in harm_results
    }

    freq_ghz = round(f0_hz / 1e9, 3)
    cpow_1 = round(cpow, 1)
    row = [freq_ghz, cpow_1] + [order2power.get(order, "") for order in order_list]

    file_exists = os.path.exists(filename)
    need_write_header = False
    rows_out: List[List[str]] = []

    if file_exists:
        try:
            rows_out = _read_table_rows(filename)

            last_header = None
            for r in reversed(rows_out):
                if len(r) >= 2 and r[0] == "Freq_GHz" and r[1] == "CPOW_dBm":
                    last_header = r
                    break

            if last_header is None:
                need_write_header = True
            else:
                existing_harm_cols = last_header[2:] if len(last_header) > 2 else []
                new_harm_cols = headers[2:]
                if existing_harm_cols != new_harm_cols:
                    need_write_header = True
                else:
                    need_write_header = False

        except Exception as e:
            print(f"[WARN] Failed to read existing result file: {e}; will write a new header.")
            need_write_header = True
    else:
        need_write_header = True

    if need_write_header:
        rows_out.append(headers)
    rows_out.append(row)

    _write_table_rows(filename, rows_out)

    print(f"\n[FILE] Wrote file: {filename}")
    print("[FILE] Wrote header:", "yes" if need_write_header else "no")
    print("[FILE] Header:", ",".join(headers))
    print("[FILE] Row:", ",".join(str(x) for x in row))


def _get_base_dir() -> Path:
    if getattr(sys, "frozen", False):
        return Path(sys.executable).parent
    return Path(__file__).parent


def _get_config_dir() -> Path:
    return _get_base_dir() / CONFIG_DIR_NAME


def _get_result_dir() -> Path:
    return _get_base_dir() / RESULT_DIR_NAME


def _get_result_bandedge_dir() -> Path:
    return _get_base_dir() / RESULT_DIR_BANDEDGE_NAME


def _get_result_bt_dir() -> Path:
    return _get_base_dir() / RESULT_DIR_BT_NAME


def _get_last_dut_path() -> Path:
    return _get_config_dir() / "last_dut_name.txt"


def _load_last_dut_name() -> str:
    try:
        path = _get_last_dut_path()
        if not path.exists():
            return ""
        return path.read_text(encoding="utf-8", errors="ignore").strip()
    except Exception:
        return ""


def _save_last_dut_name(name: str) -> None:
    cleaned = (name or "").strip()
    if not cleaned:
        return
    try:
        path = _get_last_dut_path()
        path.parent.mkdir(parents=True, exist_ok=True)
        path.write_text(cleaned, encoding="utf-8")
    except Exception:
        pass


def _append_timestamp(path: str) -> str:
    root, ext = os.path.splitext(path)
    stamp = datetime.now().strftime("%Y_%m%d_%H%M")
    return f"{root}_{stamp}{ext}"


def _resolve_config_resource(name: str) -> Path:
    base_dir = _get_base_dir()
    config_dir = base_dir / CONFIG_DIR_NAME
    candidate = config_dir / name
    if candidate.exists():
        return candidate
    fallback = base_dir / name
    if fallback.exists():
        return fallback
    bundle_root = Path(getattr(sys, "_MEIPASS", base_dir))
    candidate = bundle_root / CONFIG_DIR_NAME / name
    if candidate.exists():
        return candidate
    return bundle_root / name


def _select_profile_files(profile: str, flow_mode: str, test_type: str = "HARMONIC") -> Tuple[str, str, str]:
    mode = flow_mode.strip().upper()
    test_key = test_type.strip().upper()
    profile_key = profile.strip().upper()
    profile_loss_table = LOSS_TABLE_PATH_DULE_ANTENNA if profile_key == "DULE_ANTENNA" else LOSS_TABLE_PATH
    if mode == "BT":
        input_csv_name = INPUT_CSV_BT_BLE
        output_csv_name = OUTPUT_CSV_BT
        loss_table_name = LOSS_TABLE_PATH
        return input_csv_name, output_csv_name, loss_table_name

    if mode == "WIFI" and test_key == "BANDEDGE":
        return INPUT_CSV_BANDEDGE, OUTPUT_CSV_BANDEDGE, profile_loss_table

    input_csv_name = INPUT_CSV_SINGLE_BAND
    output_csv_name = OUTPUT_CSV
    loss_table_name = profile_loss_table
    if profile_key == "DULE_BAND":
        input_csv_name = INPUT_CSV_DULE_BAND
        output_csv_name = f"DULE_BAND_{OUTPUT_CSV}"
    elif profile_key == "DULE_ANTENNA":
        input_csv_name = INPUT_CSV_DULE_ANTENNA
        output_csv_name = f"Dule_Antenna_{OUTPUT_CSV}"
        loss_table_name = LOSS_TABLE_PATH_DULE_ANTENNA
    return input_csv_name, output_csv_name, loss_table_name


def main():
    inst = FsvSocket(FSV_IP, FSV_PORT, SOCKET_TIMEOUT)
    gui = None
    gui_proc = None
    (
        dut_name,
        profile,
        cal_scope_min,
        cal_scope_max,
        cal_scope_step,
        band_24,
        band_5,
        test_harmonic,
        test_bandedge,
        selected_connect_type,
        selected_firmware_type,
        selected_gui_override,
    ) = _prompt_user_inputs()
    selected_fw = (selected_firmware_type or "WIFI").strip().upper()
    selected_gui_override = (selected_gui_override or "").strip()
    enable_gui_log = selected_fw in {"BLE", "WIFI_AND_BLE"}
    flow_plan: List[str]
    if selected_fw == "BLE":
        flow_plan = ["BT"]
    elif selected_fw == "WIFI_AND_BLE":
        flow_plan = ["WIFI", "BT"]
    else:
        flow_plan = ["WIFI"]

    jobs: List[Dict[str, object]] = []
    for flow_mode in flow_plan:
        mode_key = flow_mode.strip().upper()
        if mode_key == "BT":
            input_csv_name, output_csv_name, loss_table_name = _select_profile_files(
                profile, flow_mode, test_type="HARMONIC"
            )
            jobs.append(
                {
                    "flow_mode": flow_mode,
                    "test_type": "HARMONIC",
                    "input_csv_name": input_csv_name,
                    "output_csv_name": output_csv_name,
                    "loss_table_name": loss_table_name,
                }
            )
            continue

        if test_harmonic:
            input_csv_name, output_csv_name, loss_table_name = _select_profile_files(
                profile, flow_mode, test_type="HARMONIC"
            )
            jobs.append(
                {
                    "flow_mode": flow_mode,
                    "test_type": "HARMONIC",
                    "input_csv_name": input_csv_name,
                    "output_csv_name": output_csv_name,
                    "loss_table_name": loss_table_name,
                }
            )
        if test_bandedge:
            input_csv_name, output_csv_name, loss_table_name = _select_profile_files(
                profile, flow_mode, test_type="BANDEDGE"
            )
            jobs.append(
                {
                    "flow_mode": flow_mode,
                    "test_type": "BANDEDGE",
                    "input_csv_name": input_csv_name,
                    "output_csv_name": output_csv_name,
                    "loss_table_name": loss_table_name,
                }
            )

    xlsx_paths: List[str] = []
    for job in jobs:
        xlsx_paths.append(str(job["input_csv_name"]))
        xlsx_paths.append(str(job["output_csv_name"]))
    if not jobs:
        print("[WARN] No test task selected. Nothing to run.")
        return
    if not _ensure_xlsx_support(xlsx_paths):
        return

    for job in jobs:
        input_path = _resolve_config_resource(str(job["input_csv_name"]))
        if not input_path.exists():
            raise FileNotFoundError(
                f"Config file not found: {input_path}. "
                f"Expected file name: {job['input_csv_name']}"
            )
        job["input_path"] = input_path
        csv_gui_host, csv_gui_port, csv_connect_type, csv_gui_exe_path = _load_global_csv_settings(
            str(input_path)
        )
        job["csv_gui_host"] = csv_gui_host
        job["csv_gui_port"] = csv_gui_port
        job["csv_connect_type"] = csv_connect_type
        job["csv_gui_exe_path"] = csv_gui_exe_path

    changed_files = _override_all_config_gui_address(selected_gui_override)
    print(f"[INFO] GUI ADDRESS override: {selected_gui_override}")
    if changed_files:
        print(f"[INFO] Updated GUI ADDRESS in {len(changed_files)} config file(s).")
        for changed in changed_files:
            print(f"[INFO]   {changed.name}")
    else:
        print("[INFO] No GUI ADDRESS field changed in config files.")
    for job in jobs:
        job["csv_gui_exe_path"] = selected_gui_override

    base_dir = _get_base_dir()
    config_dir = _get_config_dir()
    config_dir.mkdir(parents=True, exist_ok=True)
    result_dir = _get_result_dir()
    result_dir.mkdir(parents=True, exist_ok=True)
    result_bandedge_dir = _get_result_bandedge_dir()
    result_bandedge_dir.mkdir(parents=True, exist_ok=True)
    result_bt_dir = _get_result_bt_dir()
    result_bt_dir.mkdir(parents=True, exist_ok=True)
    first_job = jobs[0]
    gui_host = str(first_job.get("csv_gui_host") or GUI_HOST)
    gui_port = int(first_job.get("csv_gui_port") or GUI_PORT)
    gui_exe_path = str(first_job.get("csv_gui_exe_path") or GUI_EXE_PATH)
    local_gui_host_ip = _get_local_ipv4_for_gui()
    gui_host = local_gui_host_ip
    gui_port = 7481
    changed_host_files = _override_all_config_gui_host(gui_host)
    if changed_host_files:
        print(f"[INFO] Updated GUI HOST in {len(changed_host_files)} config file(s).")
        for changed in changed_host_files:
            print(f"[INFO]   {changed.name}")
    for job in jobs:
        job["csv_gui_host"] = gui_host
        job["csv_gui_port"] = gui_port
    print(f"[INFO] GUI HostIP (local): {gui_host}")
    print(f"[INFO] GUI Port: {gui_port}")
    written_results: List[Tuple[str, str]] = []
    if USE_GUI_CALIBRATION:
        GuiSocketClient = _load_gui_client_class()
        gui = GuiSocketClient(
            gui_host,
            gui_port,
            GUI_TIMEOUT,
            trailing_space=False,
            log_io=enable_gui_log,
        )
        if enable_gui_log:
            print("[INFO] GUI debug I/O logging enabled for BT flow.")

    try:
        if AUTO_LAUNCH_GUI and gui_exe_path:
            exe_path = Path(gui_exe_path)
            config_path = _sync_gui_main_ini(exe_path, gui_host, gui_port)
            if config_path is None:
                print(f"[WARN] GUI config not found: {exe_path.parent / 'Config' / 'main.ini'}")
            original_cwd = os.getcwd()
            os.chdir(str(exe_path.parent))
            gui_proc = subprocess.Popen([str(exe_path)])
            time.sleep(2.0)
            os.chdir(original_cwd)
        if gui:
            start_time = time.time()
            while True:
                try:
                    gui.connect()
                    break
                except Exception as exc:
                    if time.time() - start_time > 20:
                        raise
                    print(f"[WARN] GUI connect failed, retrying: {exc}")
                    time.sleep(1.0)
        inst.connect()
        idn = inst.query("*IDN?")
        print("[IDN]", idn)

        for idx, job in enumerate(jobs):
            flow_mode = str(job["flow_mode"])
            test_type = str(job.get("test_type") or "HARMONIC").strip().upper()
            input_path = Path(job["input_path"])
            output_csv_name = str(job["output_csv_name"])
            if flow_mode.upper() == "BT":
                output_base_dir = result_bt_dir
            elif test_type == "BANDEDGE":
                output_base_dir = result_bandedge_dir
            else:
                output_base_dir = result_dir
            output_path = str(output_base_dir / output_csv_name)
            if dut_name:
                output_path = str(output_base_dir / f"{dut_name}_{output_csv_name}")
            output_path = _append_timestamp(output_path)

            loss_path = _resolve_config_resource(str(job["loss_table_name"]))
            cable_loss_table = _load_cable_loss_table(str(loss_path))
            if cable_loss_table:
                print(f"[INFO] Cable loss table loaded: {len(cable_loss_table)} ranges from {loss_path}")
            else:
                print(f"[WARN] Cable loss table is empty or missing: {loss_path}")

            csv_connect_type = str(job.get("csv_connect_type") or "")
            effective_connect_type = selected_connect_type or csv_connect_type
            default_fw = "BLE" if flow_mode.upper() == "BT" else "WIFI"

            print(f"[INFO] Starting {flow_mode} {test_type} test with config: {input_path.name}")
            if flow_mode.upper() == "WIFI" and test_type == "BANDEDGE":
                run_bandedge_test(
                    input_path=str(input_path),
                    output_path=output_path,
                    inst=inst,
                    gui=gui,
                    default_connect_type=effective_connect_type,
                    default_firmware_type=default_fw,
                    cable_loss_table=cable_loss_table,
                    band_24=band_24,
                    band_5=band_5,
                    profile=profile,
                    flow_mode=flow_mode,
                )
            else:
                run_csv_test(
                    str(input_path),
                    output_path,
                    inst,
                    gui,
                    default_connect_type=effective_connect_type,
                    default_firmware_type=default_fw,
                    cable_loss_table=cable_loss_table,
                    cal_scope_min=cal_scope_min,
                    cal_scope_max=cal_scope_max,
                    cal_scope_step=cal_scope_step,
                    band_24=band_24,
                    band_5=band_5,
                    test_harmonic=True,
                    test_bandedge=False,
                    flow_mode=flow_mode,
                )
            written_results.append((output_path, str(input_path)))
            has_next_job = idx < (len(jobs) - 1)
            if has_next_job and gui:
                try:
                    print("[INFO] Intermediate GUI disconnect before next flow.")
                    gui.send("DISCONNECT")
                    _sleep_cmd(FW_SWITCH_DISCONNECT_S)
                except Exception as exc:
                    print(f"[WARN] Intermediate DISCONNECT failed: {exc}")

    finally:
        if gui:
            try:
                try:
                    gui_version = gui.get_version()
                    version_text = str(gui_version).strip() or "(empty)"
                    print(f"[RESULT] GUI GET_VERSION: {version_text}")
                    try:
                        for output_path, input_path in written_results:
                            rows = _read_table_rows(output_path) if os.path.exists(output_path) else []
                            rows.append(["GET_VERSION", version_text])
                            _write_table_rows(output_path, rows, template_path=str(input_path))
                    except Exception as exc:
                        print(f"[WARN] Failed to write GET_VERSION to result: {exc}")
                except Exception as exc:
                    print(f"[WARN] GUI GET_VERSION failed: {exc}")
                print("[GUI] DISCONNECT")
                gui.disconnect()
                _sleep_cmd(CMD_DELAY)
            except Exception:
                pass
            gui.close()
        inst.close()
        if gui_proc and gui_proc.poll() is None:
            try:
                gui_proc.terminate()
                gui_proc.wait(timeout=5)
            except Exception:
                try:
                    gui_proc.kill()
                except Exception:
                    pass
if __name__ == "__main__":
    main()
